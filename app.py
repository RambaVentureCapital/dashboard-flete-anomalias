"""
Dashboard de Auditoría Logística — v17
=======================================
Cambios respecto a v16, todos en la Pestaña 2:
  - Sección "Detalle de registros" ahora se titula según el botón
    activo:
      • Facturas auditadas      → "Detalle de registros con discrepancias"
      • Match exacto            → "Detalle de registros con match exacto"
      • Con discrepancias       → "Detalle de registros con discrepancias"
      • Errores de captura conf → "Detalle de registros con errores de captura"
  - El checkbox "Ocultar matches OK" solo aparece en el panel de
    Facturas auditadas (en los demás la data ya está pre-filtrada).
  - Sección "Casos a revisar prioritariamente":
      • Match exacto            → la sección se ELIMINA por completo.
      • Facturas auditadas      → muestra registros amarillos +
                                   naranjas + rojos (NO los verdes).
      • Con discrepancias       → muestra registros rojos + naranjas
                                   (no los verdes).
      • Errores de captura conf → mantiene comportamiento v16
                                   (registros verdes con typo).

Cambios heredados de v16:
  - Pestaña 2 (Facturas sin Match en SAP): los 4 KPIs (Facturas
    auditadas, Match exacto, Con discrepancias, Errores de captura
    confirmados) son ahora BOTONES INTERACTIVOS con la misma
    estética y dinámica que los de la Pestaña 1 (gradiente azul,
    hover, mutuamente exclusivos, fade en los inactivos).
    - Click en "Facturas auditadas" → muestra el resumen completo
      (todo el contenido que v15 enseñaba siempre).
    - Click en "Match exacto" → mismo resumen pero filtrado SOLO a
      registros con coincidencia exacta en SAP.
    - Click en "Con discrepancias" → mismo resumen filtrado a
      registros con discrepancias (excluye match y provisión).
    - Click en "Errores de captura confirmados" → mismo resumen
      filtrado SOLO a registros con typo + monto coincidente.
  - Las etiquetas de los botones permiten wrap; ya no se trunca
    "Errores de captura confirmados" con "...".

Cambios heredados de v15:
  - Las descargas de tablas ahora se generan en formato EXCEL (.xlsx)
    en lugar de CSV, preservando el formato visual de cada celda
    (colores de fondo, color de texto, negritas), tal y como se ven
    en el dashboard. CSV es texto plano y no soporta colores.
    Afectados:
      • Pestaña 1 → Panel Clientes Analizados (resumen_clientes.xlsx)
      • Pestaña 1 → Panel Anomalías ALTO (anomalias_alto.xlsx)
      • Pestaña 1 → Panel Anomalías MEDIO (anomalias_medio.xlsx)
      • Pestaña 2 → Tabla de auditoría SAP (auditoria_facturas_sap.xlsx)
  - Se re-aplica color en columnas Z_Flete, Z_Litros, Z_CxL en las
    tablas de paneles ALTO y MEDIO (rojo si |Z|>=2, amarillo si
    |Z|>=1.5) para que tanto la tabla en pantalla como el Excel
    descargado mantengan el código de colores.
  - Formatos numéricos (currency, +/-) se aplican en el Excel
    mediante openpyxl tras la escritura.

Cambios heredados de v14:
  - CORRECCIÓN visual de los botones KPI: ahora SÍ se aplica el
    fondo azul gradiente con letras blancas. El bug en v13 era que
    los selectores CSS usaban combinadores de hijo directo (>) que
    no funcionan en Streamlit 1.57 porque la clase 'st-key-{key}'
    queda en un wrapper externo separado del botón por varios
    elementos intermedios. La solución: usar selectores descendentes
    (espacio) para que coincidan con cualquier nivel de anidamiento.
  - Se aumenta la especificidad CSS con selectores múltiples para
    sobrevivir a futuros cambios del DOM de Streamlit.

Cambios heredados de v13:
  - Los 4 botones KPI de la Pestaña 1 ahora son MUTUAMENTE
    EXCLUSIVOS: solo un panel puede estar abierto a la vez. Hacer
    click en otro botón reemplaza el panel anterior por el nuevo.
    Hacer click en el mismo botón cierra el panel.
  - Indicador visual: el botón ACTIVO mantiene su color azul lleno
    con gradiente. Los botones INACTIVOS se ven desvanecidos
    (opacidad reducida y color gris) cuando hay otro abierto.

Cambios heredados de v12:
  - Se ELIMINA la duplicación de información en la Pestaña 1.
    Antes (v11) las secciones "Anomalías Detectadas", "Diagnósticos
    detallados", "Top clientes con anomalías" y "Análisis individual
    por cliente" se mostraban siempre debajo de los paneles KPI,
    duplicando lo que los paneles ya enseñaban filtrado.
  - Cada KPI ahora muestra SOLO la información que le es propia:
      • Total Flete       → distribución de Flete + Top 10 clientes
                            por Flete total + evolución mensual.
      • Clientes analizados → única y exclusivamente la tabla con
                              stats por cliente.
      • Anomalías ALTO    → Top clientes con anomalías ALTO +
                            diagnósticos detallados (ALTO) +
                            Análisis individual por cliente (limitado
                            a clientes con anomalías ALTO) + tabla
                            de registros ALTO.
      • Anomalías MEDIO   → mismas piezas que ALTO pero filtradas a
                            MEDIO (color amarillo).
  - Las secciones "siempre visibles" se eliminaron por completo: el
    contenido se mueve íntegramente dentro de los paneles KPI.

Cambios heredados de v11:
  - Tarjetas KPI clickables tipo botón con efecto hover.

Cambios heredados de v10:
  - Los valores de los KPIs nunca se truncan; escalan con el ancho.

Cambios heredados de v9:
  - La Pestaña 3 (Diagnóstico de Ventas) es ahora completamente
    INDEPENDIENTE: ya no requiere subir el Excel logístico desde la
    sidebar para desbloquearse. El usuario puede entrar directamente
    a la Pestaña 3 y usar el pipeline LLMX vs Sale_Database con sus
    propios uploaders internos.
  - Las Pestañas 1 y 2 muestran un mensaje invitando a subir el Excel
    cuando no hay archivo cargado, en lugar de bloquear toda la app.

Cambios heredados de v8:
  - PASO 3 del Pipeline filtra registros de Nacionales/Extranjeros
    por la fecha de contabilización, considerando sólo el mes anómalo
    detectado en PASO 2 (y el año del LLMX).

Dashboard con TRES pestañas:

  📈 PESTAÑA 1 — Anomalías de Flete por Cliente
     Detección estadística de saltos atípicos en el costo de flete
     usando Z-scores comparados contra el patrón histórico del cliente.

  🔎 PESTAÑA 2 — Facturas sin Match en SAP
     Auditoría que cruza las facturas de transporte de Logística Nac
     contra SAP proveedores, detectando errores de captura tipográficos,
     UUIDs inválidos y posibles duplicados.

  ⚡ PESTAÑA 3 — Diagnóstico de Ventas (Pipeline LLMX vs Sale_Database)
     Pipeline de tres pasos: limpieza C/D (ML), detección de mes
     sospechoso por discrepancia Gross Sales, e investigación de
     documentos faltantes en Nacionales/Extranjeros.

INSTALACIÓN:
    pip install streamlit pandas numpy plotly openpyxl python-Levenshtein

EJECUCIÓN:
    streamlit run app.py
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
from plotly.subplots import make_subplots
import re
import os
import io
import traceback
from collections import defaultdict
from openpyxl import load_workbook

# ============================================================
# CONFIGURACIÓN DE LA PÁGINA
# ============================================================
st.set_page_config(
    page_title="Auditoría Logística",
    page_icon="🚚",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Estilos CSS personalizados
st.markdown("""
<style>
    .main { padding-top: 1rem; }
    .stMetric {
        background: linear-gradient(135deg, #1E3A8A 0%, #1E40AF 100%);
        padding: 20px;
        border-radius: 16px;
        color: white;
        box-shadow: 0 4px 12px rgba(30, 58, 138, 0.2);
    }
    .stMetric label { color: #BFDBFE !important; font-size: 13px !important; }
    /* Evitar el truncamiento del valor del KPI: permitir wrap y nunca usar
       ellipsis. Se ajusta el tamaño con clamp() para que escale según el
       ancho disponible sin forzar overflow. */
    .stMetric [data-testid="stMetricValue"] {
        color: white !important;
        font-size: clamp(20px, 2.4vw, 34px) !important;
        line-height: 1.15 !important;
        white-space: normal !important;
        overflow: visible !important;
        text-overflow: clip !important;
        word-break: break-word !important;
        overflow-wrap: anywhere !important;
        min-width: 0 !important;
    }
    .stMetric [data-testid="stMetricValue"] > div {
        white-space: normal !important;
        overflow: visible !important;
        text-overflow: clip !important;
    }
    .stMetric [data-testid="stMetricDelta"] { color: #DBEAFE !important; }

    h1 { color: #0F172A; }
    h2 { color: #1E293B; border-bottom: 2px solid #1E3A8A; padding-bottom: 8px; }
    h3 { color: #334155; }

    .severidad-alto {
        background: linear-gradient(135deg, #DC2626 0%, #991B1B 100%);
        color: white !important;
        padding: 12px 20px;
        border-radius: 12px;
        font-weight: 600;
        margin: 4px 0;
    }
    .severidad-medio {
        background: linear-gradient(135deg, #F59E0B 0%, #B45309 100%);
        color: white !important;
        padding: 12px 20px;
        border-radius: 12px;
        font-weight: 600;
        margin: 4px 0;
    }

    /* ─── KPIs CLICKABLES (Pestañas 1 y 2) — v14/v16 fix ───── */
    /* Los botones con keys kpi_btn_* (Pestaña 1) y kpi2_btn_*
       (Pestaña 2) se estilizan como tarjetas de gradiente azul,
       idénticas visualmente a las st.metric originales.
       Usamos SELECTORES DESCENDENTES (espacio) y NO de hijo directo
       (>) porque Streamlit 1.57 inserta wrappers intermedios entre
       el elemento con clase 'st-key-{key}' y el <button>.          */
    [class*="st-key-kpi_btn_"] button,
    [class*="st-key-kpi2_btn_"] button,
    body [class*="st-key-kpi_btn_"] button,
    body [class*="st-key-kpi2_btn_"] button {
        background: linear-gradient(135deg, #1E3A8A 0%, #1E40AF 100%) !important;
        color: white !important;
        border: none !important;
        border-radius: 16px !important;
        padding: 18px 22px !important;
        min-height: 110px !important;
        height: auto !important;
        box-shadow: 0 4px 12px rgba(30, 58, 138, 0.20) !important;
        text-align: left !important;
        transition: transform 0.15s ease, box-shadow 0.15s ease, background 0.15s ease !important;
        white-space: normal !important;
        line-height: 1.3 !important;
    }
    [class*="st-key-kpi_btn_"] button:hover,
    [class*="st-key-kpi2_btn_"] button:hover,
    body [class*="st-key-kpi_btn_"] button:hover,
    body [class*="st-key-kpi2_btn_"] button:hover {
        background: linear-gradient(135deg, #2A4DC8 0%, #2A50E0 100%) !important;
        transform: translateY(-2px) !important;
        box-shadow: 0 8px 18px rgba(30, 58, 138, 0.32) !important;
        color: white !important;
        border: none !important;
    }
    [class*="st-key-kpi_btn_"] button:focus,
    [class*="st-key-kpi_btn_"] button:active,
    [class*="st-key-kpi2_btn_"] button:focus,
    [class*="st-key-kpi2_btn_"] button:active,
    body [class*="st-key-kpi_btn_"] button:focus,
    body [class*="st-key-kpi_btn_"] button:active,
    body [class*="st-key-kpi2_btn_"] button:focus,
    body [class*="st-key-kpi2_btn_"] button:active {
        background: linear-gradient(135deg, #1E3A8A 0%, #1E40AF 100%) !important;
        color: white !important;
        border: none !important;
        box-shadow: 0 4px 12px rgba(30, 58, 138, 0.20) !important;
    }
    /* Texto del botón — TODO en blanco; primer párrafo más pequeño (label),
       resto más grande (valor del KPI). */
    [class*="st-key-kpi_btn_"] button p,
    [class*="st-key-kpi_btn_"] button div,
    [class*="st-key-kpi_btn_"] button span,
    [class*="st-key-kpi2_btn_"] button p,
    [class*="st-key-kpi2_btn_"] button div,
    [class*="st-key-kpi2_btn_"] button span {
        color: white !important;
    }
    [class*="st-key-kpi_btn_"] button p:first-child,
    [class*="st-key-kpi2_btn_"] button p:first-child,
    body [class*="st-key-kpi_btn_"] button p:first-child,
    body [class*="st-key-kpi2_btn_"] button p:first-child {
        color: #BFDBFE !important;
        font-size: 13px !important;
        font-weight: 600 !important;
        margin: 0 0 4px 0 !important;
    }
    [class*="st-key-kpi_btn_"] button p:nth-child(n+2),
    [class*="st-key-kpi2_btn_"] button p:nth-child(n+2),
    body [class*="st-key-kpi_btn_"] button p:nth-child(n+2),
    body [class*="st-key-kpi2_btn_"] button p:nth-child(n+2) {
        color: white !important;
        font-size: clamp(20px, 2.4vw, 32px) !important;
        font-weight: 700 !important;
        line-height: 1.15 !important;
        margin: 4px 0 0 0 !important;
        white-space: normal !important;
        overflow: visible !important;
        text-overflow: clip !important;
        word-break: break-word !important;
    }
    /* Indicador OPEN — borde brillante cuando el panel está activo */
    div[class*="st-key-kpi_btn_"][data-kpi-open="true"] > div.stButton > button,
    div[class*="st-key-kpi_btn_"].kpi-open > div.stButton > button {
        outline: 3px solid #FBBF24 !important;
        outline-offset: 2px !important;
    }

    /* ─── PANELES DE DETALLE KPI ─────────────────────────── */
    .kpi-detail-panel {
        background: linear-gradient(180deg, #F8FAFC 0%, #F1F5F9 100%);
        border-left: 5px solid #1E3A8A;
        border-radius: 12px;
        padding: 20px 24px;
        margin: 12px 0 18px 0;
        box-shadow: 0 2px 8px rgba(30, 58, 138, 0.08);
    }
    .kpi-detail-panel h4 {
        color: #1E3A8A;
        margin-top: 0;
        margin-bottom: 14px;
        font-size: 18px;
    }
    .kpi-detail-panel.alto { border-left-color: #DC2626; background: linear-gradient(180deg, #FEF2F2 0%, #FEE2E2 100%); }
    .kpi-detail-panel.alto h4 { color: #B91C1C; }
    .kpi-detail-panel.medio { border-left-color: #F59E0B; background: linear-gradient(180deg, #FFFBEB 0%, #FEF3C7 100%); }
    .kpi-detail-panel.medio h4 { color: #B45309; }
    .kpi-detail-panel.clientes { border-left-color: #7C3AED; background: linear-gradient(180deg, #F5F3FF 0%, #EDE9FE 100%); }
    .kpi-detail-panel.clientes h4 { color: #5B21B6; }

    /* Estilos para tabs */
    .stTabs [data-baseweb="tab-list"] {
        gap: 8px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 50px;
        background: #F1F5F9;
        border-radius: 10px 10px 0 0;
        padding: 12px 20px;
        font-weight: 600;
    }
    .stTabs [aria-selected="true"] {
        background: linear-gradient(135deg, #1E3A8A 0%, #1E40AF 100%) !important;
        color: white !important;
    }
</style>
""", unsafe_allow_html=True)


# ============================================================
# UTILIDADES COMPARTIDAS
# ============================================================
@st.cache_data
def cargar_logistica(ruta_archivo):
    """Carga la hoja 'Logistica Nac' (encabezados en fila 4)."""
    df = pd.read_excel(ruta_archivo, sheet_name="Logistica Nac", header=3)
    df.columns = [str(c).strip() for c in df.columns]
    if 'Fecha Factura' in df.columns:
        df['Fecha Factura'] = pd.to_datetime(df['Fecha Factura'], errors='coerce')
    cols_numericas = ['Flete', 'Litros Fact', 'Litros Rem', 'Total Flete', 'CXL']
    for col in cols_numericas:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors='coerce')
    return df


@st.cache_data
def cargar_sap(ruta_archivo):
    """Carga la hoja 'SAP proveedores'."""
    df = pd.read_excel(ruta_archivo, sheet_name="SAP proveedores")
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ============================================================
# COLOREADORES COMPARTIDOS Y EXPORTACIÓN A EXCEL CON FORMATO (v15)
# ============================================================
def colorear_z_celda(val):
    """Color de fondo según la magnitud absoluta del Z-score."""
    try:
        v = float(val)
        if abs(v) >= 2:
            return 'background-color: #FEE2E2; color: #991B1B; font-weight: bold;'
        if abs(v) >= 1.5:
            return 'background-color: #FEF3C7; color: #92400E;'
    except (ValueError, TypeError):
        pass
    return ''


def colorear_severidad_celda(val):
    if val == 'ALTO':
        return 'background-color: #FEE2E2; color: #991B1B; font-weight: bold;'
    if val == 'MEDIO':
        return 'background-color: #FEF3C7; color: #92400E; font-weight: bold;'
    if val == 'INSUFICIENTE':
        return 'background-color: #F3F4F6; color: #6B7280; font-style: italic;'
    return ''


def styled_to_xlsx_bytes(styler, sheet_name='Datos', column_formats=None):
    """
    Convierte un pandas Styler a bytes de archivo .xlsx PRESERVANDO el formato
    de celda (background-color → fill, color → font color, font-weight: bold →
    bold). Adicionalmente aplica formatos numéricos por columna usando openpyxl
    después de escribir, ya que pandas Styler.to_excel no transfiere los
    formatos definidos con .format().

    Args:
        styler: pandas.io.formats.style.Styler ya configurado con .format() y .map().
        sheet_name: nombre de la hoja resultante.
        column_formats: dict {nombre_columna: formato_excel}, p.ej.
                         {'Flete': '"$"#,##0.00', 'Z_Flete': '+0.00;-0.00;0.00'}

    Returns:
        bytes del archivo .xlsx listo para st.download_button.
    """
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        styler.to_excel(writer, sheet_name=sheet_name, index=False)

        ws = writer.sheets[sheet_name]
        df_data = styler.data

        # Encabezado: negrita y fondo azul oscuro
        from openpyxl.styles import Font, PatternFill, Alignment
        header_font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        for col_idx in range(1, len(df_data.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        ws.row_dimensions[1].height = 28

        # Auto-ajuste de ancho aproximado por columna
        for col_idx, col_name in enumerate(df_data.columns, start=1):
            try:
                max_len = max(
                    [len(str(col_name))] +
                    [len(str(v)) for v in df_data[col_name].head(120).fillna('')]
                )
                ancho = min(max(max_len + 2, 10), 60)
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = ancho
            except Exception:
                pass

        # Formatos numéricos aplicados directamente sobre las celdas
        if column_formats:
            for col_name, fmt in column_formats.items():
                if col_name in df_data.columns:
                    col_idx = df_data.columns.get_loc(col_name) + 1
                    col_letter = ws.cell(row=1, column=col_idx).column_letter
                    for row in range(2, ws.max_row + 1):
                        cell = ws[f'{col_letter}{row}']
                        cell.number_format = fmt

        # Congelar primera fila para mejor lectura
        ws.freeze_panes = 'A2'

    return buf.getvalue()


XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


# ============================================================
# PIPELINE DIAGNÓSTICO DE VENTAS — LLMX vs Sale_Database
# (basado en "Diagnostico PNL vs Monthly v4.py")
# ============================================================
MESES_MAP = {
    'ene': 1, 'jan': 1, 'enero': 1, 'january': 1,
    'feb': 2, 'febrero': 2, 'february': 2,
    'mar': 3, 'marzo': 3, 'march': 3,
    'abr': 4, 'apr': 4, 'abril': 4, 'april': 4,
    'may': 5, 'mayo': 5,
    'jun': 6, 'junio': 6, 'june': 6,
    'jul': 7, 'julio': 7, 'july': 7,
    'ago': 8, 'aug': 8, 'agosto': 8, 'august': 8,
    'sep': 9, 'sept': 9, 'septiembre': 9, 'september': 9,
    'oct': 10, 'octubre': 10, 'october': 10,
    'nov': 11, 'noviembre': 11, 'november': 11,
    'dic': 12, 'dec': 12, 'diciembre': 12, 'december': 12,
}

SEP_PIPE = "═" * 60
SEP2_PIPE = "─" * 60

HOJAS_SALES = {'Sale_Database', 'Nacionales', 'Extranjeros'}
PATRON_PL = re.compile(r"^P&L\s+\d{2}$")
PATRON_TAB = re.compile(r"^\d+_\d{2}$")


class FormatError(Exception):
    def __init__(self, mensaje, detalles=None):
        super().__init__(mensaje)
        self.detalles = detalles or []


def _to_float_pipe(val):
    if val is None:
        return None
    try:
        return float(str(val).replace(',', '').strip())
    except (ValueError, TypeError):
        return None


def _mes_nombre_a_num(v):
    if not isinstance(v, str):
        return None
    return MESES_MAP.get(v.strip().lower())


def _extraer_anio(filename, pl_bytes):
    """Extrae año del nombre del archivo; si no, lo infiere de las hojas."""
    m = re.search(r'_(\d{4})', os.path.basename(filename or ''))
    if m:
        return int(m.group(1))
    try:
        wb_tmp = load_workbook(io.BytesIO(pl_bytes), read_only=True,
                               data_only=True, keep_links=False)
        for hoja in wb_tmp.sheetnames:
            m2 = re.match(r'^(?:P&L\s+|\d+_)(\d{2})$', hoja.strip(), re.IGNORECASE)
            if m2:
                sufijo = int(m2.group(1))
                wb_tmp.close()
                return 2000 + sufijo
        wb_tmp.close()
    except Exception:
        pass
    return None


def _detectar_fila_encabezado(ws, columnas_requeridas, max_filas=15):
    columnas_norm = {c.strip().lower(): c for c in columnas_requeridas}
    mejor_fila = None
    mejor_matches = {}
    mejor_score = 0
    for fila in ws.iter_rows(min_row=1, max_row=max_filas):
        matches = {}
        for cell in fila:
            if isinstance(cell.value, str):
                key = cell.value.strip().lower()
                if key in columnas_norm:
                    matches[columnas_norm[key]] = cell.column
        if len(matches) > mejor_score:
            mejor_score = len(matches)
            mejor_fila = fila[0].row
            mejor_matches = matches
    faltantes = [c for c in columnas_requeridas if c not in mejor_matches]
    return mejor_fila, mejor_matches, faltantes


def _detectar_cols_nYY(ws, nombre_hoja):
    col_etiqueta = None
    candidatos_valor = {}
    for fila in ws.iter_rows(min_row=2):
        for cell in fila:
            if isinstance(cell.value, str) and 'gross sales' in cell.value.strip().lower():
                col_etiqueta = cell.column
                for c in fila:
                    if c.column < col_etiqueta and _to_float_pipe(c.value) is not None:
                        candidatos_valor.setdefault(col_etiqueta, []).append(c.column)
                break
        if col_etiqueta:
            break
    if col_etiqueta is None:
        raise FormatError(f"Hoja '{nombre_hoja}': no se encontró celda 'Gross Sales'.")
    if col_etiqueta not in candidatos_valor or not candidatos_valor[col_etiqueta]:
        raise FormatError(f"Hoja '{nombre_hoja}': no hay valor numérico a la izquierda de 'Gross Sales'.")
    return col_etiqueta, max(candidatos_valor[col_etiqueta])


def _validar_columnas(cols_encontradas, cols_faltantes, nombre_hoja, fila_enc):
    if cols_faltantes:
        raise FormatError(
            f"Hoja '{nombre_hoja}': faltan columnas en fila {fila_enc}.",
            detalles=[f"No encontrada: '{c}'" for c in cols_faltantes]
            + [f"Detectadas: {list(cols_encontradas.keys())}"]
        )


def _limpiar_valor_cdml(valor):
    if not isinstance(valor, str):
        return valor
    limpio = re.sub(r'[()MXP\s]', '', valor)
    return limpio if limpio else valor


def _extraer_codigo(texto):
    if not isinstance(texto, str):
        return None
    m = re.search(r'(C\d+)\s*$', texto.strip())
    return m.group(1) if m else None


def _extraer_mes_anio_fecha(val):
    """
    Extrae (mes, año) de un valor que puede ser datetime, date o string.
    Soporta formatos comunes: DD/MM/YYYY, DD-MM-YYYY, YYYY-MM-DD, etc.
    Retorna None si no se puede parsear.
    """
    import datetime
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return (val.month, val.year)
    if isinstance(val, datetime.date):
        return (val.month, val.year)
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d',
                    '%d/%m/%y', '%d-%m-%y', '%m/%d/%Y', '%m-%d-%Y'):
            try:
                d = datetime.datetime.strptime(s, fmt)
                return (d.month, d.year)
            except ValueError:
                continue
    return None


def detectar_tipo_archivo(file_bytes):
    """Detecta si los bytes corresponden a un archivo 'sales' o 'llmx'."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True,
                           data_only=True, keep_links=False)
        hojas = set(wb.sheetnames)
        wb.close()
    except Exception:
        return None

    if HOJAS_SALES.issubset(hojas):
        return 'sales'

    tiene_pl = any(PATRON_PL.match(h) for h in hojas)
    tiene_tab = any(PATRON_TAB.match(h) for h in hojas)
    if tiene_pl and tiene_tab:
        return 'llmx'
    return None


def run_pipeline_streamlit(sale_bytes, sale_filename, pl_bytes, pl_filename, log):
    """
    Ejecuta el pipeline completo en memoria (sin tocar el disco).

    Devuelve un dict con:
      - 'cleaned_sale_bytes': bytes del Sales DB ya limpio (PASO 1)
      - 'mes_sospechoso', 'diferencia', 'ausentes', etc. para el reporte

    log(msg, tag) escribe en la consola.
    """
    def titulo(txt):
        log(f"\n{SEP_PIPE}", "sep")
        log(f"  {txt}", "titulo")
        log(SEP_PIPE, "sep")

    def ok(txt):
        log(f"  ✔  {txt}", "ok")

    def info(txt):
        log(f"  {txt}", "info")

    def warn(txt):
        log(f"  ⚠  {txt}", "warn")

    resultado = {
        'cleaned_sale_bytes': None,
        'mes_sospechoso': None,
        'diferencia': None,
        'ausentes': None,
        'mensaje_final': '',
    }

    # ── PASO 1 ──────────────────────────────────────────────────
    titulo("PASO 1 — LIMPIEZA: C/D (ML) + FACTURAS CANCELADAS")

    if not sale_bytes:
        raise FormatError("No se recibió el archivo Sales Data Base.")
    if not pl_bytes:
        raise FormatError("No se recibió el archivo LLMX.")

    wb1 = load_workbook(io.BytesIO(sale_bytes), keep_links=False)
    for nombre in ['Nacionales', 'Extranjeros']:
        if nombre not in wb1.sheetnames:
            raise FormatError(f"Hoja '{nombre}' no encontrada.",
                              [f"Hojas: {wb1.sheetnames}"])
        ws = wb1[nombre]
        fila_enc, cols, faltantes = _detectar_fila_encabezado(
            ws, ['C/D (ML)', 'Info.detallada'])
        _validar_columnas(cols, faltantes, nombre, fila_enc)
        col_cd = cols['C/D (ML)']
        col_info = cols['Info.detallada']

        limpias = 0
        for fila in ws.iter_rows(min_row=fila_enc + 1, min_col=col_cd, max_col=col_cd):
            cell = fila[0]
            if cell.value is None:
                continue
            nuevo = _limpiar_valor_cdml(str(cell.value))
            if nuevo != str(cell.value):
                cell.value = nuevo
                limpias += 1

        normales = defaultdict(list)
        cancelaciones = []
        for fila in ws.iter_rows(min_row=fila_enc + 1):
            info_val = fila[col_info - 1].value
            cd_val = fila[col_cd - 1].value
            rn = fila[0].row
            if not isinstance(info_val, str) or not info_val.strip():
                continue
            codigo = _extraer_codigo(info_val)
            try:
                monto = abs(float(str(cd_val).replace(',', '').strip())) if cd_val else None
            except (ValueError, TypeError):
                monto = None
            if not codigo or not monto:
                continue
            if 'cancelaci' in info_val.lower():
                cancelaciones.append((rn, codigo, monto))
            else:
                normales[(codigo, monto)].append(rn)

        eliminar = set()
        for rn, codigo, monto in cancelaciones:
            k = (codigo, monto)
            if k in normales and normales[k]:
                eliminar.add(rn)
                eliminar.add(normales[k].pop(0))
        for rn in sorted(eliminar, reverse=True):
            ws.delete_rows(rn)
        ok(f"{nombre}: {limpias} celdas limpias · {len(eliminar)} canceladas eliminadas")

    cleaned_buf = io.BytesIO()
    wb1.save(cleaned_buf)
    cleaned_sale_bytes = cleaned_buf.getvalue()
    resultado['cleaned_sale_bytes'] = cleaned_sale_bytes
    ok("Archivo Sales DB limpio (disponible para descarga).")

    # ── PASO 2 ──────────────────────────────────────────────────
    titulo("PASO 2 — DIAGNÓSTICO GROSS SALES")

    anio = _extraer_anio(pl_filename, pl_bytes)
    if not anio:
        raise FormatError(
            "No se detectó año en el nombre del archivo LLMX.",
            ["Formato esperado: LLMX_YYYY... (ej. LLMX_2026 v2.xlsx)"])
    sufijo = str(anio)[-2:]

    wb_pl = load_workbook(io.BytesIO(pl_bytes), data_only=True)
    nombre_hoja_pl = f"P&L {sufijo}"
    if nombre_hoja_pl not in wb_pl.sheetnames:
        raise FormatError(f"Hoja '{nombre_hoja_pl}' no encontrada.",
                          [f"Hojas disponibles: {wb_pl.sheetnames}"])
    ws_pl = wb_pl[nombre_hoja_pl]

    fila_gs = None
    for row in ws_pl.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().upper() == 'GROSS SALES':
                fila_gs = cell.row
                break
        if fila_gs:
            break
    if not fila_gs:
        raise FormatError(f"No se encontró 'GROSS SALES' en hoja '{nombre_hoja_pl}'.")

    q_cols = {}
    for fe in range(1, fila_gs):
        for cell in ws_pl[fe]:
            if isinstance(cell.value, str):
                m = re.match(r'Q(\d)[\s\-/]?\d{0,4}', cell.value.strip(), re.IGNORECASE)
                if m:
                    nq = int(m.group(1))
                    if nq not in q_cols:
                        q_cols[nq] = cell.column
    if not q_cols:
        raise FormatError(f"No se encontraron columnas Q1-Q4 en '{nombre_hoja_pl}'.")

    ultimo_q = valor_gs = None
    for q in sorted(q_cols.keys()):
        v = _to_float_pipe(ws_pl.cell(row=fila_gs, column=q_cols[q]).value)
        if v is not None and v > 0:
            ultimo_q = q
            valor_gs = v
    if ultimo_q is None:
        raise FormatError("Todos los valores Q de GROSS SALES son cero o negativos.")

    info(f"[P&L] Q{ultimo_q} detectado  |  GROSS SALES = {valor_gs:,.2f}")

    wb_sale = load_workbook(io.BytesIO(cleaned_sale_bytes), data_only=True)
    if 'Sale_Database' not in wb_sale.sheetnames:
        raise FormatError("Hoja 'Sale_Database' no encontrada en Sales Data Base.")
    ws_sale = wb_sale['Sale_Database']
    fila_enc_sale, cols_sale, faltantes = _detectar_fila_encabezado(ws_sale, ['GROSS SALES'])
    _validar_columnas(cols_sale, faltantes, 'Sale_Database', fila_enc_sale)
    col_gs_s = cols_sale['GROSS SALES']

    meses_sale = set()
    for fila in ws_sale.iter_rows(min_row=fila_enc_sale + 1, min_col=1, max_col=1):
        n = _mes_nombre_a_num(fila[0].value)
        if n:
            meses_sale.add(n)

    meses_pl = set()
    for h in wb_pl.sheetnames:
        m = re.match(rf'^(\d+)_{sufijo}$', h)
        if m:
            meses_pl.add(int(m.group(1)))

    meses_comunes = sorted(meses_sale & meses_pl)
    if not meses_comunes:
        raise FormatError(
            "Sin meses en común entre Sale_Database y tabs N_YY.",
            [f"Meses Sale_DB: {sorted(meses_sale)}",
             f"Meses LLMX:    {sorted(meses_pl)}"])

    info(f"Meses a revisar: {meses_comunes}")
    mes_sospechoso = diferencia = suma_sale_final = saldo_pl_final = None

    for mes in meses_comunes:
        suma_acum = 0.0
        for fila in ws_sale.iter_rows(min_row=fila_enc_sale + 1):
            mn = _mes_nombre_a_num(fila[0].value)
            gsv = _to_float_pipe(fila[col_gs_s - 1].value)
            if mn is not None and mn <= mes and gsv is not None and gsv > 0:
                suma_acum += gsv

        nombre_tab = f"{mes}_{sufijo}"
        if nombre_tab not in wb_pl.sheetnames:
            info(f"    Mes {mes}: tab '{nombre_tab}' no disponible.")
            continue
        ws_n = wb_pl[nombre_tab]
        col_et, col_val = _detectar_cols_nYY(ws_n, nombre_tab)
        saldo_pl = sum(
            abs(_to_float_pipe(f[col_val - 1].value))
            for f in ws_n.iter_rows(min_row=2)
            if isinstance(f[col_et - 1].value, str)
            and 'gross sales' in f[col_et - 1].value.lower()
            and _to_float_pipe(f[col_val - 1].value) is not None
        )

        coincide = round(suma_acum, 2) == round(saldo_pl, 2)
        estado = "✅" if coincide else "❌"
        info(f"    Mes {mes:>2}  Sale_DB: {suma_acum:>14,.2f}  |  LLMX: {saldo_pl:>14,.2f}  {estado}")

        if not coincide:
            mes_sospechoso = mes
            suma_sale_final = suma_acum
            saldo_pl_final = saldo_pl
            break

    if not mes_sospechoso:
        ok("Todos los meses coinciden. Sin discrepancia.")
        resultado['mensaje_final'] = "Todos los meses coinciden — sin discrepancia."
        return resultado

    diferencia = suma_sale_final - saldo_pl_final
    direccion = "MÁS" if diferencia > 0 else "MENOS"
    signo = "+" if diferencia > 0 else ""
    log(f"\n  🔍  MES SOSPECHOSO: MES {mes_sospechoso} ({anio})", "alerta")
    info(f"     Sale_DB acumulado  : {suma_sale_final:>14,.2f}")
    info(f"     LLMX tab           : {saldo_pl_final:>14,.2f}")
    info(f"     Diferencia         : {signo}{diferencia:>13,.2f}  ({abs(diferencia):,.2f} {direccion})")

    resultado['mes_sospechoso'] = mes_sospechoso
    resultado['diferencia'] = diferencia
    resultado['anio'] = anio

    # ── PASO 3 ──────────────────────────────────────────────────
    titulo(f"PASO 3 — INVESTIGACIÓN DOCUMENTOS (MES {mes_sospechoso})")

    wb3 = load_workbook(io.BytesIO(cleaned_sale_bytes), data_only=True)
    ws3 = wb3['Sale_Database']
    fila_enc3, cols3, falt3 = _detectar_fila_encabezado(ws3, ['GROSS SALES', 'Documento'])
    _validar_columnas(cols3, falt3, 'Sale_Database', fila_enc3)
    col_doc = cols3['Documento']
    col_gs3 = cols3['GROSS SALES']

    todos_docs = set()
    conteo = 0
    for fila in ws3.iter_rows(min_row=fila_enc3 + 1):
        mn = _mes_nombre_a_num(fila[0].value)
        doc = fila[col_doc - 1].value
        gsv = _to_float_pipe(fila[col_gs3 - 1].value)
        if mn == mes_sospechoso and doc is not None and gsv is not None and gsv > 0:
            todos_docs.add(str(doc).strip())
            conteo += 1

    info(f"[Sale_DB] Mes {mes_sospechoso}: {conteo} filas con GS>0  |  {len(todos_docs)} documentos distintos")

    numeros_origen = {}
    fuera_de_mes_total = 0
    sin_fecha_total = 0
    for tab_nombre in ['Nacionales', 'Extranjeros']:
        ws_t = wb3[tab_nombre]
        fila_enc_t, cols_t, falt_t = _detectar_fila_encabezado(
            ws_t, ['Número de origen', 'C/D (ML)', 'Fecha de contabilización'])
        # 'Fecha de contabilización' es altamente recomendable pero no obligatoria
        # (si falta, no podemos filtrar por mes y caemos al comportamiento v7).
        _validar_columnas(
            {k: v for k, v in cols_t.items() if k != 'Fecha de contabilización'},
            [c for c in falt_t if c != 'Fecha de contabilización'],
            tab_nombre, fila_enc_t,
        )
        col_no = cols_t['Número de origen']
        col_cd = cols_t['C/D (ML)']
        col_fecha = cols_t.get('Fecha de contabilización')

        if col_fecha is None:
            warn(f"{tab_nombre}: columna 'Fecha de contabilización' no encontrada — "
                 f"se procesarán todos los registros sin filtrar por mes.")

        contadores_local = {'incluidos': 0, 'fuera_de_mes': 0, 'sin_fecha': 0}
        for fila in ws_t.iter_rows(min_row=fila_enc_t + 1):
            val = fila[col_no - 1].value
            cd = fila[col_cd - 1].value
            if val is None:
                continue

            # Filtro por mes/año si la columna fecha existe
            if col_fecha is not None:
                fecha_val = fila[col_fecha - 1].value
                mes_anio = _extraer_mes_anio_fecha(fecha_val)
                if mes_anio is None:
                    contadores_local['sin_fecha'] += 1
                    continue
                if mes_anio != (mes_sospechoso, anio):
                    contadores_local['fuera_de_mes'] += 1
                    continue

            numeros_origen[str(val).strip()] = {
                'hoja': tab_nombre,
                'cdml': _to_float_pipe(cd),
            }
            contadores_local['incluidos'] += 1

        fuera_de_mes_total += contadores_local['fuera_de_mes']
        sin_fecha_total += contadores_local['sin_fecha']
        info(f"    {tab_nombre}: {contadores_local['incluidos']} incluidos · "
             f"{contadores_local['fuera_de_mes']} fuera del mes {mes_sospechoso}/{anio} · "
             f"{contadores_local['sin_fecha']} sin fecha legible")

    info(f"Nacionales + Extranjeros (mes {mes_sospechoso}/{anio}): "
         f"{len(numeros_origen)} registros · "
         f"omitidos por mes: {fuera_de_mes_total} · sin fecha: {sin_fecha_total}")

    ausentes = {k: v for k, v in numeros_origen.items() if k not in todos_docs}
    info(f"Presentes en Sale_DB: {len(numeros_origen) - len(ausentes)}  |  Ausentes: {len(ausentes)}")

    if ausentes:
        log(f"\n  {SEP2_PIPE}", "sep")
        log(f"  FACTURAS EN N/E NO ENCONTRADAS EN Sale_DB:", "alerta")
        log(f"  {SEP2_PIPE}", "sep")
        log(f"  {'Origen':<15} {'Hoja':<14} {'C/D (ML)':>14}", "header")
        log(f"  {'-' * 15} {'-' * 14} {'-' * 14}", "sep")
        total = 0.0
        for origen in sorted(ausentes.keys()):
            d = ausentes[origen]
            m_val = d['cdml'] or 0.0
            total += m_val
            log(f"  {origen:<15} {d['hoja']:<14} {m_val:>14,.2f}", "fila")
        log(f"  {SEP2_PIPE}", "sep")
        log(f"  {'TOTAL':<30} {total:>14,.2f}", "total")
        log(f"  {'Diferencia esperada':<30} {abs(diferencia):>14,.2f}", "total")
        if round(abs(total), 2) == round(abs(diferencia), 2):
            log(f"\n  ✅  Las facturas ausentes EXPLICAN la diferencia.", "ok")
            resultado['mensaje_final'] = "Las facturas ausentes explican la diferencia."
        else:
            brecha = abs(diferencia) - abs(total)
            warn(f"Queda una brecha de {brecha:,.2f} sin explicar.")
            resultado['mensaje_final'] = f"Brecha sin explicar: {brecha:,.2f}"
        resultado['ausentes'] = ausentes
    else:
        ok("Todos los registros de N/E están presentes en Sale_Database.")
        resultado['mensaje_final'] = "Todos los registros N/E presentes en Sale_DB."

    return resultado


# ============================================================
# DETECCIÓN DE ANOMALÍAS POR CLIENTE
# ============================================================
def detectar_anomalias_por_cliente(df, umbral_z_alto=2.0, umbral_z_medio=1.5, min_registros=3):
    """Detecta anomalías comparando contra la media histórica del MISMO CLIENTE."""
    df = df.copy()
    df = df.dropna(subset=['Fecha Factura', 'Nombre de Cliente'])
    df = df[df['Flete'].notna() & (df['Flete'] > 0)]
    df = df[df['Litros Fact'].notna() & (df['Litros Fact'] > 0)]
    df['Costo por Litro'] = df['Flete'] / df['Litros Fact']

    stats_cliente = df.groupby('Nombre de Cliente').agg(
        Flete_media=('Flete', 'mean'),
        Flete_std=('Flete', 'std'),
        Litros_media=('Litros Fact', 'mean'),
        Litros_std=('Litros Fact', 'std'),
        CxL_media=('Costo por Litro', 'mean'),
        CxL_std=('Costo por Litro', 'std'),
        Num_registros=('Flete', 'count')
    ).reset_index()

    df = df.merge(stats_cliente, on='Nombre de Cliente', how='left')

    df['Z_Flete'] = np.where(df['Flete_std'] > 0,
                             (df['Flete'] - df['Flete_media']) / df['Flete_std'], 0)
    df['Z_Litros'] = np.where(df['Litros_std'] > 0,
                              (df['Litros Fact'] - df['Litros_media']) / df['Litros_std'], 0)
    df['Z_CxL'] = np.where(df['CxL_std'] > 0,
                           (df['Costo por Litro'] - df['CxL_media']) / df['CxL_std'], 0)

    def clasificar(row):
        if row['Num_registros'] < min_registros:
            return 'INSUFICIENTE'
        z_flete = row['Z_Flete']
        z_litros = row['Z_Litros']
        z_cxl = row['Z_CxL']
        if z_cxl < umbral_z_medio:
            return 'NORMAL'
        if z_cxl >= umbral_z_alto and (z_flete >= umbral_z_alto or z_litros < 0.5):
            return 'ALTO'
        if z_cxl >= umbral_z_medio:
            return 'MEDIO'
        return 'NORMAL'

    df['Severidad'] = df.apply(clasificar, axis=1)
    df['Es Anomalía'] = df['Severidad'].isin(['ALTO', 'MEDIO']).astype(int)

    def diagnostico(row):
        if row['Severidad'] == 'NORMAL':
            return ''
        if row['Severidad'] == 'INSUFICIENTE':
            return f'⚪ Cliente con pocos registros ({int(row["Num_registros"])}) - no evaluable'
        z_flete, z_litros, z_cxl = row['Z_Flete'], row['Z_Litros'], row['Z_CxL']
        sf = '+' if z_flete > 0 else ''
        sl = '+' if z_litros > 0 else ''
        sc = '+' if z_cxl > 0 else ''
        if row['Severidad'] == 'ALTO':
            if abs(z_litros) < 1:
                return (f'🔴 FLETE atípico ({sf}{z_flete:.1f}σ vs media cliente '
                        f'${row["Flete_media"]:,.0f}) SIN salto proporcional en LITROS '
                        f'({sl}{z_litros:.1f}σ). CxL ${row["Costo por Litro"]:.2f} '
                        f'vs media ${row["CxL_media"]:.2f} ({sc}{z_cxl:.1f}σ). '
                        f'Probable error o cargo no justificado — REVISAR.')
            return (f'🔴 FLETE atípico ({sf}{z_flete:.1f}σ) y CxL elevado '
                    f'({sc}{z_cxl:.1f}σ vs media cliente). REVISAR.')
        return (f'🟡 FLETE alto ({sf}{z_flete:.1f}σ) y litros '
                f'{"también altos" if abs(z_litros) > 0.5 else "normales"} '
                f'({sl}{z_litros:.1f}σ); CxL elevado ({sc}{z_cxl:.1f}σ). Verificar.')

    df['Diagnóstico'] = df.apply(diagnostico, axis=1)
    return df


# ============================================================
# AUDITORÍA: FACTURAS SIN MATCH EN SAP
# ============================================================
UUID_REGEX = re.compile(r'^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}$')


def es_uuid(texto):
    """Verifica si una cadena tiene formato UUID válido."""
    if not isinstance(texto, str):
        return False
    return bool(UUID_REGEX.match(texto.strip()))


def parece_uuid(texto):
    """UUID 'casi válido' (~36 chars con guiones) — útil para typos."""
    if not isinstance(texto, str):
        return False
    t = texto.strip()
    if len(t) < 30 or len(t) > 42:
        return False
    return t.count('-') >= 3


def levenshtein(s1, s2):
    """Distancia de Levenshtein entre dos cadenas (implementación pura)."""
    if not isinstance(s1, str) or not isinstance(s2, str):
        return 999
    s1, s2 = s1.upper().strip(), s2.upper().strip()
    if len(s1) < len(s2):
        s1, s2 = s2, s1
    if len(s2) == 0:
        return len(s1)
    previous = list(range(len(s2) + 1))
    for i, c1 in enumerate(s1):
        current = [i + 1]
        for j, c2 in enumerate(s2):
            insert_cost = previous[j + 1] + 1
            delete_cost = current[j] + 1
            replace_cost = previous[j] + (c1 != c2)
            current.append(min(insert_cost, delete_cost, replace_cost))
        previous = current
    return previous[-1]


def explicar_typo(texto_logistica, texto_sap):
    """Genera explicación humana del tipo de error tipográfico."""
    a = str(texto_logistica).upper().strip()
    b = str(texto_sap).upper().strip()
    if a == b:
        return "coincidencia exacta"
    if len(a) != len(b):
        diff = abs(len(a) - len(b))
        if diff == 1:
            return f"falta o sobra 1 carácter"
        return f"diferencia de {diff} caracteres en longitud"
    # Misma longitud — buscar diferencias
    diferencias = []
    for i, (ca, cb) in enumerate(zip(a, b)):
        if ca != cb:
            diferencias.append((i, ca, cb))
    if len(diferencias) == 1:
        i, ca, cb = diferencias[0]
        return f"un carácter cambiado en posición {i + 1}: '{ca}' debería ser '{cb}'"
    if len(diferencias) == 2:
        # Verificar si son caracteres intercambiados
        d1, d2 = diferencias
        if d1[1] == d2[2] and d1[2] == d2[1] and d2[0] - d1[0] <= 2:
            return f"caracteres intercambiados ('{d1[1]}' y '{d2[1]}' aparecen invertidos)"
        return f"2 caracteres distintos"
    return f"{len(diferencias)} caracteres distintos"


def split_facturas(texto):
    """Divide una celda con varias facturas separadas por /, , o ;."""
    if pd.isna(texto):
        return []
    return [t.strip() for t in re.split(r'[/,;]', str(texto)) if t.strip()]


def auditar_facturas(df_log, df_sap, tolerancia_monto=1.0):
    """
    Cruza Logística Nac contra SAP proveedores y clasifica cada factura.

    Categorías:
    - PROVISION (provisión contable, no requiere match)
    - COINCIDE_EXACTO (match en SAP, sin issues)
    - UUID_TYPO_CON_MONTO_COINCIDENTE / UUID_POSIBLE_TYPO_SIN_MONTO
    - UUID_MONTO_COINCIDE_DISTINTO / UUID_NO_ENCONTRADO / UUID_FORMATO_INVALIDO
    - TYPO_CON_MONTO_COINCIDENTE / POSIBLE_TYPO_SIN_MONTO
    - MONTO_COINCIDE_REF_DISTINTA / NO_ENCONTRADO
    """
    # Preparar SAP: referencias, UUIDs y montos
    sap_refs = df_sap['Referencia Factura'].dropna().astype(str).str.strip().tolist()
    sap_uuids = df_sap['UUID Factura'].dropna().astype(str).str.strip().tolist()

    # Map: ref/uuid -> (índice fila SAP, total sin IVA, total movimiento)
    ref_to_data = {}
    for idx, row in df_sap.iterrows():
        ref = row.get('Referencia Factura')
        if pd.notna(ref):
            ref_to_data[str(ref).strip().upper()] = {
                'fila_sap': idx + 2,  # +2 por el header de Excel (no 0-index)
                'total_sin_iva': row.get('Total Sin IVA'),
                'total_mov': row.get('Total Movimiento'),
                'ref_original': str(ref).strip()
            }

    uuid_to_data = {}
    for idx, row in df_sap.iterrows():
        u = row.get('UUID Factura')
        if pd.notna(u):
            uuid_to_data[str(u).strip().upper()] = {
                'fila_sap': idx + 2,
                'total_sin_iva': row.get('Total Sin IVA'),
                'total_mov': row.get('Total Movimiento'),
                'uuid_original': str(u).strip()
            }

    resultados = []

    for idx, row in df_log.iterrows():
        fac_transp = row.get('Factura transporte')
        proveedor = row.get('Proveedor transporte')
        total_flete = row.get('Total Flete') if 'Total Flete' in row.index else row.get('Total Flete ')

        # Filtros: ignorar filas vacías o proveedor = CLIENTE
        if pd.isna(fac_transp):
            continue
        if pd.notna(proveedor) and str(proveedor).strip().upper() == 'CLIENTE':
            continue

        # Una celda puede traer varias facturas separadas
        facturas = split_facturas(fac_transp)
        if not facturas:
            continue

        for factura in facturas:
            f_upper = factura.upper().strip()

            # PROVISION
            if 'PROVISION' in f_upper or 'PROVISON' in f_upper:
                resultados.append({
                    'idx_log': idx,
                    'factura_evaluada': factura,
                    'Categoría': 'PROVISION',
                    'Razón': 'Marca de provisión contable — no requiere match en SAP.',
                })
                continue

            # ¿Es UUID?
            es_uuid_valido = es_uuid(factura)
            parece_uuid_aprox = parece_uuid(factura) and not es_uuid_valido

            if es_uuid_valido or parece_uuid_aprox:
                # ===== CASO UUID =====
                if not es_uuid_valido:
                    # Formato no estándar
                    resultados.append({
                        'idx_log': idx,
                        'factura_evaluada': factura,
                        'Categoría': 'UUID_FORMATO_INVALIDO',
                        'Razón': f'El identificador "{factura}" parece UUID pero su formato es inválido (debe ser 8-4-4-4-12 caracteres hexadecimales).',
                    })
                    continue

                # Match exacto en UUIDs SAP
                if f_upper in uuid_to_data:
                    data = uuid_to_data[f_upper]
                    resultados.append({
                        'idx_log': idx,
                        'factura_evaluada': factura,
                        'Categoría': 'COINCIDE_EXACTO',
                        'Razón': f'UUID encontrado en SAP (fila {data["fila_sap"]}, monto ${data["total_sin_iva"]:,.2f}).',
                    })
                    continue

                # Buscar UUIDs similares (Levenshtein <= 3)
                candidatos = []
                for u_sap, data in uuid_to_data.items():
                    d = levenshtein(f_upper, u_sap)
                    if d <= 3 and d > 0:
                        candidatos.append((d, u_sap, data))
                candidatos.sort()

                if candidatos:
                    # ¿Algún candidato tiene monto coincidente?
                    match_monto = None
                    for d, u_sap, data in candidatos:
                        for col_monto in ['total_sin_iva', 'total_mov']:
                            if pd.notna(data[col_monto]) and pd.notna(total_flete):
                                if abs(float(data[col_monto]) - float(total_flete)) <= tolerancia_monto:
                                    match_monto = (d, u_sap, data, col_monto)
                                    break
                        if match_monto:
                            break

                    if match_monto:
                        d, u_sap, data, col = match_monto
                        explicacion = explicar_typo(factura, data['uuid_original'])
                        resultados.append({
                            'idx_log': idx,
                            'factura_evaluada': factura,
                            'Categoría': 'UUID_TYPO_CON_MONTO_COINCIDENTE',
                            'Razón': (f'POSIBLE ERROR DE CAPTURA en UUID: "{factura}" en Logística vs '
                                      f'"{data["uuid_original"]}" en SAP (fila {data["fila_sap"]}). '
                                      f'Monto coincide (${total_flete:,.2f}). Diferencia: {explicacion}. '
                                      f'Alta confianza de error tipográfico al copiar/pegar.'),
                        })
                    else:
                        candidatos_str = '; '.join([
                            f'"{c[1]}" (fila {c[2]["fila_sap"]}, ${c[2]["total_sin_iva"]:,.2f})'
                            for c in candidatos[:3]
                        ])
                        resultados.append({
                            'idx_log': idx,
                            'factura_evaluada': factura,
                            'Categoría': 'UUID_POSIBLE_TYPO_SIN_MONTO',
                            'Razón': (f'UUID "{factura}" no encontrado. Candidatos similares: {candidatos_str}. '
                                      f'Verificar manualmente.'),
                        })
                    continue

                # Sin similares — ¿algún UUID con monto exacto?
                if pd.notna(total_flete):
                    for u_sap, data in uuid_to_data.items():
                        for col_monto in ['total_sin_iva', 'total_mov']:
                            if pd.notna(data[col_monto]):
                                if abs(float(data[col_monto]) - float(total_flete)) <= tolerancia_monto:
                                    resultados.append({
                                        'idx_log': idx,
                                        'factura_evaluada': factura,
                                        'Categoría': 'UUID_MONTO_COINCIDE_DISTINTO',
                                        'Razón': (f'UUID "{factura}" no encontrado, pero hay un UUID en SAP con '
                                                  f'monto exacto ${total_flete:,.2f}: "{data["uuid_original"]}" '
                                                  f'(fila {data["fila_sap"]}). Posible factura registrada con UUID diferente.'),
                                    })
                                    break
                        else:
                            continue
                        break
                    else:
                        resultados.append({
                            'idx_log': idx,
                            'factura_evaluada': factura,
                            'Categoría': 'UUID_NO_ENCONTRADO',
                            'Razón': f'UUID "{factura}" no existe en SAP y no hay coincidencias por monto.',
                        })
                else:
                    resultados.append({
                        'idx_log': idx,
                        'factura_evaluada': factura,
                        'Categoría': 'UUID_NO_ENCONTRADO',
                        'Razón': f'UUID "{factura}" no existe en SAP.',
                    })

            else:
                # ===== CASO REFERENCIA (texto/clave/número) =====
                # Match exacto en referencias
                if f_upper in ref_to_data:
                    data = ref_to_data[f_upper]
                    resultados.append({
                        'idx_log': idx,
                        'factura_evaluada': factura,
                        'Categoría': 'COINCIDE_EXACTO',
                        'Razón': f'Referencia encontrada en SAP (fila {data["fila_sap"]}, monto ${data["total_sin_iva"]:,.2f}).',
                    })
                    continue

                # Buscar similares (Levenshtein <= 2)
                candidatos = []
                for r_sap, data in ref_to_data.items():
                    d = levenshtein(f_upper, r_sap)
                    if 0 < d <= 2:
                        candidatos.append((d, r_sap, data))
                candidatos.sort()

                if candidatos:
                    # ¿Monto coincidente?
                    match_monto = None
                    for d, r_sap, data in candidatos:
                        for col_monto in ['total_sin_iva', 'total_mov']:
                            if pd.notna(data[col_monto]) and pd.notna(total_flete):
                                if abs(float(data[col_monto]) - float(total_flete)) <= tolerancia_monto:
                                    match_monto = (d, r_sap, data, col_monto)
                                    break
                        if match_monto:
                            break

                    if match_monto:
                        d, r_sap, data, col = match_monto
                        explicacion = explicar_typo(factura, data['ref_original'])
                        resultados.append({
                            'idx_log': idx,
                            'factura_evaluada': factura,
                            'Categoría': 'TYPO_CON_MONTO_COINCIDENTE',
                            'Razón': (f'POSIBLE ERROR DE CAPTURA en referencia: "{factura}" en Logística vs '
                                      f'"{data["ref_original"]}" en SAP (fila {data["fila_sap"]}). '
                                      f'Monto coincide (${total_flete:,.2f}). Diferencia: {explicacion}. '
                                      f'Probable error tipográfico — se sugiere corregir en Logística.'),
                        })
                    else:
                        candidatos_str = '; '.join([
                            f'"{c[2]["ref_original"]}" (fila {c[2]["fila_sap"]}, ${c[2]["total_sin_iva"]:,.2f})'
                            for c in candidatos[:3]
                        ])
                        resultados.append({
                            'idx_log': idx,
                            'factura_evaluada': factura,
                            'Categoría': 'POSIBLE_TYPO_SIN_MONTO',
                            'Razón': (f'Referencia "{factura}" no encontrada. Candidatos similares: {candidatos_str}. '
                                      f'Verificar manualmente.'),
                        })
                    continue

                # Sin similares — ¿monto exacto en otra ref?
                if pd.notna(total_flete):
                    encontrado = False
                    for r_sap, data in ref_to_data.items():
                        for col_monto in ['total_sin_iva', 'total_mov']:
                            if pd.notna(data[col_monto]):
                                if abs(float(data[col_monto]) - float(total_flete)) <= tolerancia_monto:
                                    resultados.append({
                                        'idx_log': idx,
                                        'factura_evaluada': factura,
                                        'Categoría': 'MONTO_COINCIDE_REF_DISTINTA',
                                        'Razón': (f'Referencia "{factura}" no encontrada, pero hay una factura en SAP '
                                                  f'con monto exacto ${total_flete:,.2f}: "{data["ref_original"]}" '
                                                  f'(fila {data["fila_sap"]}). Posible factura registrada con clave diferente.'),
                                    })
                                    encontrado = True
                                    break
                        if encontrado:
                            break
                    if not encontrado:
                        resultados.append({
                            'idx_log': idx,
                            'factura_evaluada': factura,
                            'Categoría': 'NO_ENCONTRADO',
                            'Razón': f'Referencia "{factura}" no existe en SAP y no hay coincidencias por monto.',
                        })
                else:
                    resultados.append({
                        'idx_log': idx,
                        'factura_evaluada': factura,
                        'Categoría': 'NO_ENCONTRADO',
                        'Razón': f'Referencia "{factura}" no existe en SAP.',
                    })

    # Construir DataFrame de resultados con todas las columnas de Logística + categoría/razón
    if not resultados:
        return pd.DataFrame()

    df_res = pd.DataFrame(resultados)
    df_log_full = df_log.reset_index(drop=False).rename(columns={'index': 'idx_log_orig'})
    df_log_full['idx_log_orig'] = df_log_full.index  # asegurar que los idx coincidan

    # Reset index del df_log original para hacer merge
    df_log_indexed = df_log.reset_index(drop=True)
    df_log_indexed['idx_log'] = df_log_indexed.index

    df_final = df_res.merge(df_log_indexed, on='idx_log', how='left')

    # Renombrar para humano
    if 'Categoría' in df_final.columns and 'Razón' in df_final.columns:
        df_final = df_final.rename(columns={'Razón': 'Razón de No Coincidencia'})

    return df_final


# ============================================================
# DICCIONARIO DE CATEGORÍAS — Etiquetas humanas y colores
# ============================================================
CATEGORIAS_INFO = {
    'COINCIDE_EXACTO': {
        'label': '✅ Coincide exactamente',
        'color': '#16A34A',
        'bg': '#DCFCE7',
        'desc': 'La factura existe en SAP con coincidencia exacta. Sin observaciones.',
    },
    'PROVISION': {
        'label': '📋 Provisión contable',
        'color': '#3B82F6',
        'bg': '#DBEAFE',
        'desc': 'Asiento de provisión contable — no requiere match con factura física.',
    },
    'TYPO_CON_MONTO_COINCIDENTE': {
        'label': '🟢 Error de captura confirmado (monto coincide)',
        'color': '#15803D',
        'bg': '#BBF7D0',
        'desc': 'La referencia en Logística difiere de SAP por un error de captura, pero el monto coincide exactamente. Alta confianza de que es un error tipográfico al transcribir.',
    },
    'UUID_TYPO_CON_MONTO_COINCIDENTE': {
        'label': '🟢 Error de captura en UUID (monto coincide)',
        'color': '#15803D',
        'bg': '#BBF7D0',
        'desc': 'El UUID en Logística difiere ligeramente del de SAP, pero el monto coincide. Probable error al copiar/pegar el UUID.',
    },
    'POSIBLE_TYPO_SIN_MONTO': {
        'label': '🟠 Posible error de captura (monto NO verificable)',
        'color': '#C2410C',
        'bg': '#FED7AA',
        'desc': 'Hay referencias parecidas en SAP pero el monto no coincide o no se pudo verificar. Requiere revisión manual.',
    },
    'UUID_POSIBLE_TYPO_SIN_MONTO': {
        'label': '🟠 Posible error de captura en UUID',
        'color': '#C2410C',
        'bg': '#FED7AA',
        'desc': 'Hay UUIDs parecidos en SAP pero el monto no coincide. Requiere revisión manual.',
    },
    'MONTO_COINCIDE_REF_DISTINTA': {
        'label': '🟡 Monto coincide pero referencia distinta',
        'color': '#A16207',
        'bg': '#FEF08A',
        'desc': 'La referencia no existe en SAP, pero hay una factura con el mismo monto registrada con una clave totalmente diferente.',
    },
    'UUID_MONTO_COINCIDE_DISTINTO': {
        'label': '🟡 Monto coincide pero UUID distinto',
        'color': '#A16207',
        'bg': '#FEF08A',
        'desc': 'El UUID no existe en SAP, pero hay una factura con el mismo monto bajo otro UUID.',
    },
    'NO_ENCONTRADO': {
        'label': '🔴 No encontrado en SAP',
        'color': '#B91C1C',
        'bg': '#FECACA',
        'desc': 'La referencia no existe en SAP y no hay coincidencias por monto. Posible factura faltante de registrar.',
    },
    'UUID_NO_ENCONTRADO': {
        'label': '🔴 UUID no encontrado en SAP',
        'color': '#B91C1C',
        'bg': '#FECACA',
        'desc': 'El UUID no existe en SAP y no hay coincidencias por monto.',
    },
    'UUID_FORMATO_INVALIDO': {
        'label': '⚪ UUID con formato inválido',
        'color': '#525252',
        'bg': '#E5E5E5',
        'desc': 'El identificador parece UUID pero no cumple el formato 8-4-4-4-12 hexadecimal.',
    },
}

# ============================================================
# SIDEBAR — CARGA DE ARCHIVO
# ============================================================
st.sidebar.title("⚙️ Configuración")

archivo = st.sidebar.file_uploader(
    "📁 Sube tu archivo Excel",
    type=['xlsx', 'xls'],
    help="Excel con hojas 'Logistica Nac' y 'SAP proveedores'"
)

# ============================================================
# HEADER (siempre visible)
# ============================================================
st.title("🚚 Dashboard de Auditoría Logística")

# Cargar ambas hojas SI hay archivo; si no, dejar None y permitir
# que la Pestaña 3 funcione de forma independiente.
df_log_raw = None
df_sap = None

if archivo is None:
    st.info("📤 Sube el archivo Excel desde la barra lateral para activar las "
            "**Pestañas 1 y 2**. La **Pestaña 3** (Diagnóstico de Ventas) "
            "funciona de forma independiente — abre la pestaña directamente.")

    col_a, col_b, col_c = st.columns(3)
    with col_a:
        st.markdown("""
        ### 📈 Pestaña 1 — Anomalías de Flete
        Detecta saltos atípicos en el costo del transporte comparando
        cada operación contra el patrón histórico **del mismo cliente**.

        - Análisis estadístico con Z-scores
        - Tres niveles de severidad
        - Visualización con bandas estadísticas
        """)
    with col_b:
        st.markdown("""
        ### 🔎 Pestaña 2 — Facturas vs SAP
        Cruza las facturas de transporte de Logística contra SAP
        proveedores y detecta discrepancias.

        - Detección de errores de captura
        - Búsqueda por monto cuando falla la referencia
        - Diagnóstico humano en cada caso
        """)
    with col_c:
        st.markdown("""
        ### ⚡ Pestaña 3 — Diagnóstico de Ventas
        Pipeline LLMX vs Sale_Database para detectar mes sospechoso
        y facturas faltantes.

        - Limpieza C/D (ML) y facturas canceladas
        - Detección automática de mes con discrepancia
        - Investigación de documentos faltantes en N/E
        """)
else:
    st.markdown("Sistema integrado de detección de anomalías y conciliación de facturas")
    try:
        df_log_raw = cargar_logistica(archivo)
    except Exception as e:
        st.error(f"Error al cargar 'Logistica Nac': {e}")
        df_log_raw = None
    try:
        df_sap = cargar_sap(archivo)
    except Exception as e:
        st.error(f"Error al cargar 'SAP proveedores': {e}")
        df_sap = None

# ============================================================
# TABS
# ============================================================
tab1, tab2, tab3 = st.tabs([
    "📈  Anomalías de Flete por Cliente",
    "🔎  Facturas sin Match en SAP",
    "⚡  Diagnóstico de Ventas"
])

# ╔══════════════════════════════════════════════════════════╗
# ║  PESTAÑA 1 — ANOMALÍAS DE FLETE                          ║
# ╚══════════════════════════════════════════════════════════╝
with tab1:
    if not (df_log_raw is not None):
        st.info("📤 Sube el archivo Excel desde la barra lateral para activar esta pestaña.")

if df_log_raw is not None:
    with tab1:
        st.markdown("**Análisis estadístico individualizado** — cada cliente se compara contra su propio patrón histórico")

        # Alerta de alcance
        st.markdown("""
        <div style="
            background: linear-gradient(135deg, #FEF3C7 0%, #FDE68A 100%);
            border-left: 6px solid #F59E0B;
            padding: 16px 20px;
            border-radius: 12px;
            margin: 16px 0;
            box-shadow: 0 2px 4px rgba(245, 158, 11, 0.1);
        ">
            <div style="font-size: 16px; font-weight: 600; color: #78350F; margin-bottom: 8px;">
                ⚠️ Alcance del análisis — Lee antes de interpretar
            </div>
            <div style="color: #78350F; line-height: 1.6;">
                Este análisis usa únicamente la columna <strong>"Flete"</strong> de la hoja 
                <em>Logística Nac</em>, que corresponde al <strong>costo base del transporte</strong>. 
                <strong>NO incluye</strong> cargos adicionales como Custodia, Estadías/Repartos, 
                Permisos, Rebate (NC), FEE Logístico ni Flete Lukoil. Estos forman parte del 
                <strong>Total Flete</strong> pero se omiten aquí para no introducir ruido en la 
                detección de anomalías de tarifa base.
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Parámetros sidebar (solo aplicables a tab1)
        st.sidebar.markdown("---")
        st.sidebar.markdown("### 📈 Pestaña: Anomalías de Flete")
        umbral_alto = st.sidebar.slider("Umbral ALTO (σ)", 1.5, 3.5, 2.0, 0.1)
        umbral_medio = st.sidebar.slider("Umbral MEDIO (σ)", 1.0, 2.5, 1.5, 0.1)
        min_registros = st.sidebar.slider("Mínimo registros por cliente", 2, 10, 3)

        df = detectar_anomalias_por_cliente(df_log_raw, umbral_alto, umbral_medio, min_registros)

        # Filtros
        st.sidebar.markdown("**Filtros (Pestaña 1)**")
        clientes = sorted(df['Nombre de Cliente'].dropna().unique())
        cliente_sel = st.sidebar.multiselect("Cliente", options=clientes, default=[], key="cli_t1")

        fecha_min = df['Fecha Factura'].min().date()
        fecha_max = df['Fecha Factura'].max().date()
        rango_fechas = st.sidebar.date_input(
            "Rango de fechas", value=(fecha_min, fecha_max),
            min_value=fecha_min, max_value=fecha_max, key="fec_t1"
        )

        severidad_sel = st.sidebar.multiselect(
            "Severidad",
            options=['ALTO', 'MEDIO', 'NORMAL', 'INSUFICIENTE'],
            default=[], key="sev_t1"
        )

        # Aplicar filtros
        df_filtrado = df.copy()
        if cliente_sel:
            df_filtrado = df_filtrado[df_filtrado['Nombre de Cliente'].isin(cliente_sel)]
        if isinstance(rango_fechas, tuple) and len(rango_fechas) == 2:
            df_filtrado = df_filtrado[
                (df_filtrado['Fecha Factura'].dt.date >= rango_fechas[0]) &
                (df_filtrado['Fecha Factura'].dt.date <= rango_fechas[1])
                ]
        if severidad_sel:
            df_filtrado = df_filtrado[df_filtrado['Severidad'].isin(severidad_sel)]

        # ─── KPIs CLICKABLES (v13: mutuamente exclusivos) ──
        # Estado único: 'kpi_active' = None | 'total' | 'clientes' | 'alto' | 'medio'
        if 'kpi_active' not in st.session_state:
            st.session_state.kpi_active = None

        def _toggle_kpi(name):
            """Toggle: si ya está activo cierra; si no, lo activa (cerrando otros)."""
            st.session_state.kpi_active = None if st.session_state.kpi_active == name else name

        # Calcular valores
        kpi_total_flete = df_filtrado['Flete'].sum()
        kpi_clientes = df_filtrado['Nombre de Cliente'].nunique()
        kpi_alto = int((df_filtrado['Severidad'] == 'ALTO').sum())
        kpi_medio = int((df_filtrado['Severidad'] == 'MEDIO').sum())

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            _arr = "▲" if st.session_state.kpi_active == 'total' else "▼"
            if st.button(
                f"💰 Total Flete\n\n${kpi_total_flete:,.0f}  {_arr}",
                key="kpi_btn_total",
                use_container_width=True,
                help="Click para ver detalle (distribución, top clientes, evolución mensual)",
            ):
                _toggle_kpi('total')
        with col2:
            _arr = "▲" if st.session_state.kpi_active == 'clientes' else "▼"
            if st.button(
                f"👥 Clientes analizados\n\n{kpi_clientes}  {_arr}",
                key="kpi_btn_clientes",
                use_container_width=True,
                help="Click para ver tabla por cliente",
            ):
                _toggle_kpi('clientes')
        with col3:
            _arr = "▲" if st.session_state.kpi_active == 'alto' else "▼"
            if st.button(
                f"🔴 Anomalías ALTO\n\n{kpi_alto}  {_arr}",
                key="kpi_btn_alto",
                use_container_width=True,
                help="Click para ver detalle de anomalías ALTO",
            ):
                _toggle_kpi('alto')
        with col4:
            _arr = "▲" if st.session_state.kpi_active == 'medio' else "▼"
            if st.button(
                f"🟡 Anomalías MEDIO\n\n{kpi_medio}  {_arr}",
                key="kpi_btn_medio",
                use_container_width=True,
                help="Click para ver detalle de anomalías MEDIO",
            ):
                _toggle_kpi('medio')

        # ─── CSS dinámico: si hay un KPI activo, desvanecer los demás ──
        # Usa selectores descendentes para coincidir con el DOM de Streamlit 1.57
        # (ver nota en el CSS estático arriba).
        _active = st.session_state.kpi_active
        if _active is not None:
            _all = ('total', 'clientes', 'alto', 'medio')
            _inactive = [k for k in _all if k != _active]
            _fade_css = ""
            for _k in _inactive:
                _fade_css += (
                    f'[class*="st-key-kpi_btn_{_k}"] button,'
                    f'body [class*="st-key-kpi_btn_{_k}"] button {{'
                    f'  opacity: 0.40 !important;'
                    f'  background: linear-gradient(135deg, #94A3B8 0%, #64748B 100%) !important;'
                    f'  box-shadow: none !important;'
                    f'  transform: none !important;'
                    f'}}'
                    f'[class*="st-key-kpi_btn_{_k}"] button:hover,'
                    f'body [class*="st-key-kpi_btn_{_k}"] button:hover {{'
                    f'  opacity: 0.70 !important;'
                    f'  background: linear-gradient(135deg, #64748B 0%, #475569 100%) !important;'
                    f'  transform: none !important;'
                    f'  box-shadow: 0 2px 6px rgba(100, 116, 139, 0.25) !important;'
                    f'}}'
                    f'[class*="st-key-kpi_btn_{_k}"] button p,'
                    f'body [class*="st-key-kpi_btn_{_k}"] button p {{'
                    f'  color: #E2E8F0 !important;'
                    f'}}'
                )
            st.markdown(f"<style>{_fade_css}</style>", unsafe_allow_html=True)

        # ─── PANELES DE DETALLE — visibles según estado ─────
        # Panel: Total Flete
        if st.session_state.kpi_active == 'total':
            st.markdown('<div class="kpi-detail-panel"><h4>💰 Detalle: Total Flete</h4></div>',
                        unsafe_allow_html=True)
            n_reg = len(df_filtrado)
            if n_reg > 0:
                # Distribución (histograma)
                fig_hist = go.Figure(go.Histogram(
                    x=df_filtrado['Flete'], nbinsx=40,
                    marker_color='#1E3A8A',
                    hovertemplate='Rango: $%{x:,.0f}<br>Operaciones: %{y}<extra></extra>'
                ))
                fig_hist.update_layout(
                    title='Distribución de Flete',
                    height=300, plot_bgcolor='white', paper_bgcolor='white',
                    margin=dict(l=20, r=20, t=40, b=30),
                    xaxis=dict(title='Flete ($)'), yaxis=dict(title='# Operaciones')
                )
                st.plotly_chart(fig_hist, use_container_width=True, key="kpi_chart_hist_flete")

                # Top 10 clientes por Flete total
                top_cli = (df_filtrado.groupby('Nombre de Cliente')['Flete'].sum()
                           .sort_values(ascending=False).head(10).reset_index())
                if len(top_cli) > 0:
                    fig_top = go.Figure(go.Bar(
                        y=top_cli['Nombre de Cliente'], x=top_cli['Flete'],
                        orientation='h', marker_color='#1E40AF',
                        text=[f'${v:,.0f}' for v in top_cli['Flete']],
                        textposition='outside',
                        hovertemplate='<b>%{y}</b><br>Flete: $%{x:,.2f}<extra></extra>',
                    ))
                    fig_top.update_layout(
                        title='Top 10 clientes por Flete total',
                        height=max(280, len(top_cli) * 32),
                        plot_bgcolor='white', paper_bgcolor='white',
                        margin=dict(l=20, r=80, t=40, b=20),
                        yaxis=dict(autorange='reversed'),
                        xaxis=dict(title='Flete total ($)')
                    )
                    st.plotly_chart(fig_top, use_container_width=True, key="kpi_chart_top_flete")

                # Evolución mensual del Flete
                if 'Fecha Factura' in df_filtrado.columns:
                    df_m = df_filtrado.dropna(subset=['Fecha Factura']).copy()
                    if len(df_m) > 0:
                        df_m['Mes'] = df_m['Fecha Factura'].dt.to_period('M').astype(str)
                        monthly = df_m.groupby('Mes', as_index=False)['Flete'].sum().sort_values('Mes')
                        fig_m = go.Figure(go.Scatter(
                            x=monthly['Mes'], y=monthly['Flete'],
                            mode='lines+markers',
                            line=dict(color='#1E3A8A', width=2.5),
                            marker=dict(size=8, color='#1E40AF'),
                            hovertemplate='<b>%{x}</b><br>Flete: $%{y:,.0f}<extra></extra>',
                        ))
                        fig_m.update_layout(
                            title='Evolución mensual del Flete',
                            height=280, plot_bgcolor='white', paper_bgcolor='white',
                            margin=dict(l=20, r=20, t=40, b=20),
                            yaxis=dict(title='Flete total ($)')
                        )
                        st.plotly_chart(fig_m, use_container_width=True, key="kpi_chart_monthly_flete")
            else:
                st.info("Sin registros bajo los filtros actuales.")

        # Panel: Clientes analizados
        if st.session_state.kpi_active == 'clientes':
            st.markdown('<div class="kpi-detail-panel clientes"><h4>👥 Detalle: Clientes Analizados</h4></div>',
                        unsafe_allow_html=True)
            if df_filtrado['Nombre de Cliente'].notna().any():
                resumen_cli = df_filtrado.groupby('Nombre de Cliente').agg(
                    Registros=('Flete', 'count'),
                    Flete_Total=('Flete', 'sum'),
                    Flete_Promedio=('Flete', 'mean'),
                    CxL_Promedio=('Costo por Litro', 'mean'),
                    Anomalias_ALTO=('Severidad', lambda x: int((x == 'ALTO').sum())),
                    Anomalias_MEDIO=('Severidad', lambda x: int((x == 'MEDIO').sum())),
                ).reset_index().sort_values('Flete_Total', ascending=False)

                fmt_cli = {
                    'Flete_Total': '${:,.0f}',
                    'Flete_Promedio': '${:,.0f}',
                    'CxL_Promedio': '${:.2f}',
                }
                fmt_cli = {k: v for k, v in fmt_cli.items() if k in resumen_cli.columns}

                # Resaltar las columnas de anomalías para hacer la tabla más llamativa
                def _color_anom_alto(v):
                    try:
                        if int(v) > 0:
                            return 'background-color: #FEE2E2; color: #991B1B; font-weight: bold;'
                    except (ValueError, TypeError):
                        pass
                    return ''

                def _color_anom_medio(v):
                    try:
                        if int(v) > 0:
                            return 'background-color: #FEF3C7; color: #92400E; font-weight: bold;'
                    except (ValueError, TypeError):
                        pass
                    return ''

                styled_cli = resumen_cli.style.format(fmt_cli)
                if 'Anomalias_ALTO' in resumen_cli.columns:
                    styled_cli = styled_cli.map(_color_anom_alto, subset=['Anomalias_ALTO'])
                if 'Anomalias_MEDIO' in resumen_cli.columns:
                    styled_cli = styled_cli.map(_color_anom_medio, subset=['Anomalias_MEDIO'])

                st.caption(f"Mostrando **{len(resumen_cli)}** clientes — ordenados por Flete Total. "
                           "Las columnas son ordenables haciendo click en el encabezado.")
                st.dataframe(styled_cli, use_container_width=True, height=420)

                xlsx_cli = styled_to_xlsx_bytes(
                    styled_cli,
                    sheet_name='Resumen por cliente',
                    column_formats={
                        'Registros': '#,##0',
                        'Flete_Total': '"$"#,##0',
                        'Flete_Promedio': '"$"#,##0',
                        'CxL_Promedio': '"$"#,##0.00',
                        'Anomalias_ALTO': '0',
                        'Anomalias_MEDIO': '0',
                    }
                )
                st.download_button(
                    "📥 Descargar resumen por cliente (Excel con formato)",
                    xlsx_cli, "resumen_clientes.xlsx", XLSX_MIME,
                    key="dl_kpi_clientes"
                )
            else:
                st.info("Sin clientes bajo los filtros actuales.")

        # Panel: Anomalías ALTO
        if st.session_state.kpi_active == 'alto':
            st.markdown('<div class="kpi-detail-panel alto"><h4>🔴 Detalle: Anomalías ALTO</h4></div>',
                        unsafe_allow_html=True)
            df_alto = df_filtrado[df_filtrado['Severidad'] == 'ALTO'].copy()
            if len(df_alto) > 0:
                # ─── Top clientes con anomalías ALTO ──────────
                st.markdown("##### 📊 Top clientes con anomalías ALTO")
                cli_alto = df_alto.groupby('Nombre de Cliente').agg(
                    Anomalias=('Es Anomalía', 'sum'),
                    Flete_Total=('Flete', 'sum')
                ).reset_index().sort_values('Anomalias', ascending=False).head(15)

                fig_a_top = go.Figure(go.Bar(
                    y=cli_alto['Nombre de Cliente'], x=cli_alto['Anomalias'],
                    orientation='h', marker_color='#DC2626',
                    text=cli_alto['Anomalias'], textposition='outside',
                    hovertemplate='<b>%{y}</b><br>Anomalías ALTO: %{x}<br>Flete Total: $%{customdata:,.0f}<extra></extra>',
                    customdata=cli_alto['Flete_Total']
                ))
                fig_a_top.update_layout(
                    height=max(280, len(cli_alto) * 32),
                    plot_bgcolor='white', paper_bgcolor='white',
                    margin=dict(l=20, r=60, t=20, b=20),
                    yaxis=dict(autorange='reversed'),
                    xaxis=dict(title='Cantidad de anomalías ALTO')
                )
                st.plotly_chart(fig_a_top, use_container_width=True, key="kpi_chart_top_alto")

                # ─── Diagnósticos detallados (solo ALTO) ──────
                st.markdown("##### 🔍 Diagnósticos detallados (ALTO)")
                df_a_diag = df_alto.sort_values('Z_Flete', key=lambda x: x.abs(), ascending=False)
                for _, row in df_a_diag.iterrows():
                    fecha_str = row['Fecha Factura'].strftime("%d/%m/%Y") if pd.notna(row['Fecha Factura']) else "N/A"
                    st.markdown(
                        f'<div class="severidad-alto"><strong>{row["Nombre de Cliente"]}</strong> · '
                        f'{fecha_str} · Remisión {row.get("Remisión", "N/A")}<br>'
                        f'<small>{row["Diagnóstico"]}</small></div>',
                        unsafe_allow_html=True
                    )

                # ─── Análisis individual (limitado a clientes con ALTO) ──
                st.markdown("##### 📈 Análisis individual por cliente (con anomalías ALTO)")
                cli_disp_alto = sorted(df_alto['Nombre de Cliente'].dropna().unique().tolist())
                if cli_disp_alto:
                    cli_det_a = st.selectbox(
                        "Selecciona un cliente con anomalías ALTO",
                        options=cli_disp_alto,
                        key="sel_cli_alto"
                    )
                    df_c = df[df['Nombre de Cliente'] == cli_det_a].sort_values('Fecha Factura')
                    if len(df_c) > 0:
                        ca, cb, cc, cd = st.columns(4)
                        ca.metric("Registros", f"{len(df_c)}")
                        cb.metric("Flete promedio", f"${df_c['Flete'].mean():,.0f}")
                        cc.metric("CxL promedio", f"${df_c['Costo por Litro'].mean():.2f}")
                        cd.metric("Anomalías", f"{int(df_c['Es Anomalía'].sum())}")

                        fig_ia = make_subplots(
                            rows=3, cols=1, shared_xaxes=True, vertical_spacing=0.08,
                            subplot_titles=('Flete pagado', 'Litros facturados',
                                            'Costo por Litro (con bandas ±1σ y ±2σ)'),
                            row_heights=[0.32, 0.32, 0.36]
                        )
                        es_a = df_c['Es Anomalía'] == 1
                        cl_a = ['#DC2626' if a else '#1E3A8A' for a in es_a]
                        sz_a = [12 if a else 6 for a in es_a]

                        fig_ia.add_trace(go.Scatter(
                            x=df_c['Fecha Factura'], y=df_c['Flete'], mode='lines+markers',
                            line=dict(color='#1E3A8A', width=2),
                            marker=dict(size=sz_a, color=cl_a),
                            customdata=df_c[['Remisión', 'Folio NC', 'Proveedor transporte']].fillna('N/A').values,
                            hovertemplate=('<b>%{x|%d/%m/%Y}</b><br>Flete: $%{y:,.0f}<br>'
                                           'Remisión: %{customdata[0]}<br>Folio NC: %{customdata[1]}<br>'
                                           'Proveedor: %{customdata[2]}<extra></extra>'),
                            showlegend=False
                        ), row=1, col=1)
                        mf = df_c['Flete'].mean()
                        fig_ia.add_hline(y=mf, line_dash="dash", line_color="#888780",
                                         annotation_text=f"Media: ${mf:,.0f}", annotation_position="right",
                                         row=1, col=1)

                        fig_ia.add_trace(go.Scatter(
                            x=df_c['Fecha Factura'], y=df_c['Litros Fact'], mode='lines+markers',
                            line=dict(color='#10B981', width=2), marker=dict(size=6, color='#10B981'),
                            customdata=df_c[['Remisión', 'Folio NC', 'Proveedor transporte']].fillna('N/A').values,
                            hovertemplate=('<b>%{x|%d/%m/%Y}</b><br>Litros: %{y:,.0f}<br>'
                                           'Remisión: %{customdata[0]}<br>Folio NC: %{customdata[1]}<br>'
                                           'Proveedor: %{customdata[2]}<extra></extra>'),
                            showlegend=False
                        ), row=2, col=1)
                        ml = df_c['Litros Fact'].mean()
                        fig_ia.add_hline(y=ml, line_dash="dash", line_color="#888780",
                                         annotation_text=f"Media: {ml:,.0f}", annotation_position="right",
                                         row=2, col=1)

                        mc = df_c['Costo por Litro'].mean()
                        sc = df_c['Costo por Litro'].std()
                        b2s = mc + 2 * sc
                        b2i = max(mc - 2 * sc, 0)
                        b1s = mc + sc
                        b1i = max(mc - sc, 0)
                        fechas = df_c['Fecha Factura'].tolist()

                        fig_ia.add_trace(go.Scatter(x=fechas, y=[b2s] * len(fechas), mode='lines',
                                                    line=dict(color='#DC2626', width=1, dash='dot'),
                                                    hovertemplate=f'+2σ: ${b2s:.2f}<extra></extra>',
                                                    showlegend=False), row=3, col=1)
                        fig_ia.add_trace(go.Scatter(x=fechas, y=[b2i] * len(fechas), mode='lines',
                                                    line=dict(color='#DC2626', width=1, dash='dot'),
                                                    fill='tonexty', fillcolor='rgba(220, 38, 38, 0.05)',
                                                    hovertemplate=f'-2σ: ${b2i:.2f}<extra></extra>',
                                                    showlegend=False), row=3, col=1)
                        fig_ia.add_trace(go.Scatter(x=fechas, y=[b1s] * len(fechas), mode='lines',
                                                    line=dict(color='#F59E0B', width=1, dash='dash'),
                                                    hovertemplate=f'+1σ: ${b1s:.2f}<extra></extra>',
                                                    showlegend=False), row=3, col=1)
                        fig_ia.add_trace(go.Scatter(x=fechas, y=[b1i] * len(fechas), mode='lines',
                                                    line=dict(color='#F59E0B', width=1, dash='dash'),
                                                    fill='tonexty', fillcolor='rgba(16, 185, 129, 0.08)',
                                                    hovertemplate=f'-1σ: ${b1i:.2f}<extra></extra>',
                                                    showlegend=False), row=3, col=1)
                        fig_ia.add_hline(y=mc, line_dash="dash", line_color="#475569",
                                         annotation_text=f"Media: ${mc:.2f}", annotation_position="right",
                                         row=3, col=1)

                        cxl_v = df_c['Costo por Litro'].values
                        cl_c = []
                        sz_c = []
                        for v in cxl_v:
                            if v > b2s or v < b2i:
                                cl_c.append('#DC2626'); sz_c.append(12)
                            elif v > b1s or v < b1i:
                                cl_c.append('#F59E0B'); sz_c.append(8)
                            else:
                                cl_c.append('#1E3A8A'); sz_c.append(6)

                        fig_ia.add_trace(go.Scatter(
                            x=df_c['Fecha Factura'], y=df_c['Costo por Litro'], mode='lines+markers',
                            line=dict(color='#1E3A8A', width=2.5),
                            marker=dict(size=sz_c, color=cl_c, line=dict(color=cl_c, width=1)),
                            customdata=df_c[['Remisión', 'Folio NC', 'Flete', 'Litros Fact',
                                             'Proveedor transporte']].fillna('N/A').values,
                            hovertemplate=('<b>%{x|%d/%m/%Y}</b><br>$/L: $%{y:.2f}<br>'
                                           'Remisión: %{customdata[0]}<br>Folio NC: %{customdata[1]}<br>'
                                           'Flete: $%{customdata[2]:,.0f}<br>Litros: %{customdata[3]:,.0f}<br>'
                                           'Proveedor: %{customdata[4]}<extra></extra>'),
                            showlegend=False
                        ), row=3, col=1)

                        fig_ia.update_layout(height=700, showlegend=False,
                                             plot_bgcolor='white', paper_bgcolor='white',
                                             margin=dict(l=20, r=80, t=40, b=20), hovermode='x unified')
                        fig_ia.update_xaxes(showgrid=False)
                        fig_ia.update_yaxes(showgrid=True, gridcolor='#F1F5F9')
                        fig_ia.update_yaxes(title_text='$', row=1, col=1)
                        fig_ia.update_yaxes(title_text='L', row=2, col=1)
                        fig_ia.update_yaxes(title_text='$/L', row=3, col=1)
                        st.plotly_chart(fig_ia, use_container_width=True, key="kpi_chart_ind_alto")

                        cl1, cl2, cl3 = st.columns(3)
                        cl1.markdown('<div style="background:rgba(16,185,129,0.15);padding:8px 12px;'
                                     'border-radius:8px;border-left:4px solid #10B981;font-size:13px;">'
                                     '🟢 <strong>Zona normal</strong> (±1σ): comportamiento esperado</div>',
                                     unsafe_allow_html=True)
                        cl2.markdown('<div style="background:rgba(245,158,11,0.15);padding:8px 12px;'
                                     'border-radius:8px;border-left:4px solid #F59E0B;font-size:13px;">'
                                     '🟡 <strong>Zona de alerta</strong> (1σ–2σ): elevado pero no crítico</div>',
                                     unsafe_allow_html=True)
                        cl3.markdown('<div style="background:rgba(220,38,38,0.15);padding:8px 12px;'
                                     'border-radius:8px;border-left:4px solid #DC2626;font-size:13px;">'
                                     '🔴 <strong>Zona crítica</strong> (>2σ): anomalía estadística</div>',
                                     unsafe_allow_html=True)

                # ─── Tabla detallada ALTO + descarga Excel ────
                st.markdown(f"##### 📋 Registros con severidad ALTO ({len(df_alto)})")
                cols_a = ['Fecha Factura', 'Nombre de Cliente', 'Remisión', 'Folio NC',
                         'Proveedor transporte', 'Flete', 'Z_Flete',
                         'Litros Fact', 'Z_Litros', 'Costo por Litro', 'Z_CxL', 'Diagnóstico']
                cols_a = [c for c in cols_a if c in df_alto.columns]
                df_a_view = df_alto[cols_a].sort_values('Z_Flete', ascending=False) if 'Z_Flete' in df_alto.columns else df_alto[cols_a]
                df_a_view = df_a_view.copy()
                if 'Fecha Factura' in df_a_view.columns:
                    df_a_view['Fecha Factura'] = df_a_view['Fecha Factura'].dt.strftime('%d/%m/%Y')

                fmt_a = {
                    'Flete': '${:,.2f}', 'Litros Fact': '{:,.0f}',
                    'Costo por Litro': '${:,.2f}',
                    'Z_Flete': '{:+.2f}', 'Z_Litros': '{:+.2f}', 'Z_CxL': '{:+.2f}',
                }
                fmt_a = {k: v for k, v in fmt_a.items() if k in df_a_view.columns}

                # Estilizado: colores en Z-scores
                styled_a = df_a_view.style.format(fmt_a)
                for _zcol in ('Z_Flete', 'Z_Litros', 'Z_CxL'):
                    if _zcol in df_a_view.columns:
                        styled_a = styled_a.map(colorear_z_celda, subset=[_zcol])
                st.dataframe(styled_a, use_container_width=True, height=380)

                # Excel con formato preservado (colores + número)
                xlsx_a = styled_to_xlsx_bytes(
                    styled_a,
                    sheet_name='Anomalías ALTO',
                    column_formats={
                        'Flete': '"$"#,##0.00',
                        'Litros Fact': '#,##0',
                        'Costo por Litro': '"$"#,##0.00',
                        'Z_Flete': '+0.00;-0.00;0.00',
                        'Z_Litros': '+0.00;-0.00;0.00',
                        'Z_CxL': '+0.00;-0.00;0.00',
                    }
                )
                st.download_button(
                    "📥 Descargar registros ALTO (Excel con formato)",
                    xlsx_a, "anomalias_alto.xlsx", XLSX_MIME,
                    key="dl_kpi_alto"
                )
            else:
                st.success("✅ No hay anomalías de severidad ALTO con los filtros actuales.")

        # Panel: Anomalías MEDIO
        if st.session_state.kpi_active == 'medio':
            st.markdown('<div class="kpi-detail-panel medio"><h4>🟡 Detalle: Anomalías MEDIO</h4></div>',
                        unsafe_allow_html=True)
            df_medio = df_filtrado[df_filtrado['Severidad'] == 'MEDIO'].copy()
            if len(df_medio) > 0:
                # ─── Top clientes con anomalías MEDIO ─────────
                st.markdown("##### 📊 Top clientes con anomalías MEDIO")
                cli_medio = df_medio.groupby('Nombre de Cliente').agg(
                    Anomalias=('Es Anomalía', 'sum'),
                    Flete_Total=('Flete', 'sum')
                ).reset_index().sort_values('Anomalias', ascending=False).head(15)

                fig_m_top = go.Figure(go.Bar(
                    y=cli_medio['Nombre de Cliente'], x=cli_medio['Anomalias'],
                    orientation='h', marker_color='#F59E0B',
                    text=cli_medio['Anomalias'], textposition='outside',
                    hovertemplate='<b>%{y}</b><br>Anomalías MEDIO: %{x}<br>Flete Total: $%{customdata:,.0f}<extra></extra>',
                    customdata=cli_medio['Flete_Total']
                ))
                fig_m_top.update_layout(
                    height=max(280, len(cli_medio) * 32),
                    plot_bgcolor='white', paper_bgcolor='white',
                    margin=dict(l=20, r=60, t=20, b=20),
                    yaxis=dict(autorange='reversed'),
                    xaxis=dict(title='Cantidad de anomalías MEDIO')
                )
                st.plotly_chart(fig_m_top, use_container_width=True, key="kpi_chart_top_medio")

                # ─── Diagnósticos detallados (solo MEDIO) ─────
                st.markdown("##### 🔍 Diagnósticos detallados (MEDIO)")
                df_m_diag = df_medio.sort_values('Z_Flete', key=lambda x: x.abs(), ascending=False)
                for _, row in df_m_diag.iterrows():
                    fecha_str = row['Fecha Factura'].strftime("%d/%m/%Y") if pd.notna(row['Fecha Factura']) else "N/A"
                    st.markdown(
                        f'<div class="severidad-medio"><strong>{row["Nombre de Cliente"]}</strong> · '
                        f'{fecha_str} · Remisión {row.get("Remisión", "N/A")}<br>'
                        f'<small>{row["Diagnóstico"]}</small></div>',
                        unsafe_allow_html=True
                    )

                # ─── Análisis individual (limitado a clientes con MEDIO) ──
                st.markdown("##### 📈 Análisis individual por cliente (con anomalías MEDIO)")
                cli_disp_medio = sorted(df_medio['Nombre de Cliente'].dropna().unique().tolist())
                if cli_disp_medio:
                    cli_det_m = st.selectbox(
                        "Selecciona un cliente con anomalías MEDIO",
                        options=cli_disp_medio,
                        key="sel_cli_medio"
                    )
                    df_c = df[df['Nombre de Cliente'] == cli_det_m].sort_values('Fecha Factura')
                    if len(df_c) > 0:
                        ca, cb, cc, cd = st.columns(4)
                        ca.metric("Registros", f"{len(df_c)}")
                        cb.metric("Flete promedio", f"${df_c['Flete'].mean():,.0f}")
                        cc.metric("CxL promedio", f"${df_c['Costo por Litro'].mean():.2f}")
                        cd.metric("Anomalías", f"{int(df_c['Es Anomalía'].sum())}")

                        fig_im = make_subplots(
                            rows=3, cols=1, shared_xaxes=True, vertical_spacing=0.08,
                            subplot_titles=('Flete pagado', 'Litros facturados',
                                            'Costo por Litro (con bandas ±1σ y ±2σ)'),
                            row_heights=[0.32, 0.32, 0.36]
                        )
                        es_a = df_c['Es Anomalía'] == 1
                        cl_a = ['#DC2626' if a else '#1E3A8A' for a in es_a]
                        sz_a = [12 if a else 6 for a in es_a]

                        fig_im.add_trace(go.Scatter(
                            x=df_c['Fecha Factura'], y=df_c['Flete'], mode='lines+markers',
                            line=dict(color='#1E3A8A', width=2),
                            marker=dict(size=sz_a, color=cl_a),
                            customdata=df_c[['Remisión', 'Folio NC', 'Proveedor transporte']].fillna('N/A').values,
                            hovertemplate=('<b>%{x|%d/%m/%Y}</b><br>Flete: $%{y:,.0f}<br>'
                                           'Remisión: %{customdata[0]}<br>Folio NC: %{customdata[1]}<br>'
                                           'Proveedor: %{customdata[2]}<extra></extra>'),
                            showlegend=False
                        ), row=1, col=1)
                        mf = df_c['Flete'].mean()
                        fig_im.add_hline(y=mf, line_dash="dash", line_color="#888780",
                                         annotation_text=f"Media: ${mf:,.0f}", annotation_position="right",
                                         row=1, col=1)

                        fig_im.add_trace(go.Scatter(
                            x=df_c['Fecha Factura'], y=df_c['Litros Fact'], mode='lines+markers',
                            line=dict(color='#10B981', width=2), marker=dict(size=6, color='#10B981'),
                            customdata=df_c[['Remisión', 'Folio NC', 'Proveedor transporte']].fillna('N/A').values,
                            hovertemplate=('<b>%{x|%d/%m/%Y}</b><br>Litros: %{y:,.0f}<br>'
                                           'Remisión: %{customdata[0]}<br>Folio NC: %{customdata[1]}<br>'
                                           'Proveedor: %{customdata[2]}<extra></extra>'),
                            showlegend=False
                        ), row=2, col=1)
                        ml = df_c['Litros Fact'].mean()
                        fig_im.add_hline(y=ml, line_dash="dash", line_color="#888780",
                                         annotation_text=f"Media: {ml:,.0f}", annotation_position="right",
                                         row=2, col=1)

                        mc = df_c['Costo por Litro'].mean()
                        sc = df_c['Costo por Litro'].std()
                        b2s = mc + 2 * sc
                        b2i = max(mc - 2 * sc, 0)
                        b1s = mc + sc
                        b1i = max(mc - sc, 0)
                        fechas = df_c['Fecha Factura'].tolist()

                        fig_im.add_trace(go.Scatter(x=fechas, y=[b2s] * len(fechas), mode='lines',
                                                    line=dict(color='#DC2626', width=1, dash='dot'),
                                                    hovertemplate=f'+2σ: ${b2s:.2f}<extra></extra>',
                                                    showlegend=False), row=3, col=1)
                        fig_im.add_trace(go.Scatter(x=fechas, y=[b2i] * len(fechas), mode='lines',
                                                    line=dict(color='#DC2626', width=1, dash='dot'),
                                                    fill='tonexty', fillcolor='rgba(220, 38, 38, 0.05)',
                                                    hovertemplate=f'-2σ: ${b2i:.2f}<extra></extra>',
                                                    showlegend=False), row=3, col=1)
                        fig_im.add_trace(go.Scatter(x=fechas, y=[b1s] * len(fechas), mode='lines',
                                                    line=dict(color='#F59E0B', width=1, dash='dash'),
                                                    hovertemplate=f'+1σ: ${b1s:.2f}<extra></extra>',
                                                    showlegend=False), row=3, col=1)
                        fig_im.add_trace(go.Scatter(x=fechas, y=[b1i] * len(fechas), mode='lines',
                                                    line=dict(color='#F59E0B', width=1, dash='dash'),
                                                    fill='tonexty', fillcolor='rgba(16, 185, 129, 0.08)',
                                                    hovertemplate=f'-1σ: ${b1i:.2f}<extra></extra>',
                                                    showlegend=False), row=3, col=1)
                        fig_im.add_hline(y=mc, line_dash="dash", line_color="#475569",
                                         annotation_text=f"Media: ${mc:.2f}", annotation_position="right",
                                         row=3, col=1)

                        cxl_v = df_c['Costo por Litro'].values
                        cl_c = []
                        sz_c = []
                        for v in cxl_v:
                            if v > b2s or v < b2i:
                                cl_c.append('#DC2626'); sz_c.append(12)
                            elif v > b1s or v < b1i:
                                cl_c.append('#F59E0B'); sz_c.append(8)
                            else:
                                cl_c.append('#1E3A8A'); sz_c.append(6)

                        fig_im.add_trace(go.Scatter(
                            x=df_c['Fecha Factura'], y=df_c['Costo por Litro'], mode='lines+markers',
                            line=dict(color='#1E3A8A', width=2.5),
                            marker=dict(size=sz_c, color=cl_c, line=dict(color=cl_c, width=1)),
                            customdata=df_c[['Remisión', 'Folio NC', 'Flete', 'Litros Fact',
                                             'Proveedor transporte']].fillna('N/A').values,
                            hovertemplate=('<b>%{x|%d/%m/%Y}</b><br>$/L: $%{y:.2f}<br>'
                                           'Remisión: %{customdata[0]}<br>Folio NC: %{customdata[1]}<br>'
                                           'Flete: $%{customdata[2]:,.0f}<br>Litros: %{customdata[3]:,.0f}<br>'
                                           'Proveedor: %{customdata[4]}<extra></extra>'),
                            showlegend=False
                        ), row=3, col=1)

                        fig_im.update_layout(height=700, showlegend=False,
                                             plot_bgcolor='white', paper_bgcolor='white',
                                             margin=dict(l=20, r=80, t=40, b=20), hovermode='x unified')
                        fig_im.update_xaxes(showgrid=False)
                        fig_im.update_yaxes(showgrid=True, gridcolor='#F1F5F9')
                        fig_im.update_yaxes(title_text='$', row=1, col=1)
                        fig_im.update_yaxes(title_text='L', row=2, col=1)
                        fig_im.update_yaxes(title_text='$/L', row=3, col=1)
                        st.plotly_chart(fig_im, use_container_width=True, key="kpi_chart_ind_medio")

                        cl1, cl2, cl3 = st.columns(3)
                        cl1.markdown('<div style="background:rgba(16,185,129,0.15);padding:8px 12px;'
                                     'border-radius:8px;border-left:4px solid #10B981;font-size:13px;">'
                                     '🟢 <strong>Zona normal</strong> (±1σ): comportamiento esperado</div>',
                                     unsafe_allow_html=True)
                        cl2.markdown('<div style="background:rgba(245,158,11,0.15);padding:8px 12px;'
                                     'border-radius:8px;border-left:4px solid #F59E0B;font-size:13px;">'
                                     '🟡 <strong>Zona de alerta</strong> (1σ–2σ): elevado pero no crítico</div>',
                                     unsafe_allow_html=True)
                        cl3.markdown('<div style="background:rgba(220,38,38,0.15);padding:8px 12px;'
                                     'border-radius:8px;border-left:4px solid #DC2626;font-size:13px;">'
                                     '🔴 <strong>Zona crítica</strong> (>2σ): anomalía estadística</div>',
                                     unsafe_allow_html=True)

                # ─── Tabla detallada MEDIO + descarga Excel ───
                st.markdown(f"##### 📋 Registros con severidad MEDIO ({len(df_medio)})")
                cols_m = ['Fecha Factura', 'Nombre de Cliente', 'Remisión', 'Folio NC',
                         'Proveedor transporte', 'Flete', 'Z_Flete',
                         'Litros Fact', 'Z_Litros', 'Costo por Litro', 'Z_CxL', 'Diagnóstico']
                cols_m = [c for c in cols_m if c in df_medio.columns]
                df_m_view = df_medio[cols_m].sort_values('Z_Flete', ascending=False) if 'Z_Flete' in df_medio.columns else df_medio[cols_m]
                df_m_view = df_m_view.copy()
                if 'Fecha Factura' in df_m_view.columns:
                    df_m_view['Fecha Factura'] = df_m_view['Fecha Factura'].dt.strftime('%d/%m/%Y')

                fmt_m = {
                    'Flete': '${:,.2f}', 'Litros Fact': '{:,.0f}',
                    'Costo por Litro': '${:,.2f}',
                    'Z_Flete': '{:+.2f}', 'Z_Litros': '{:+.2f}', 'Z_CxL': '{:+.2f}',
                }
                fmt_m = {k: v for k, v in fmt_m.items() if k in df_m_view.columns}

                # Estilizado: colores en Z-scores
                styled_m = df_m_view.style.format(fmt_m)
                for _zcol in ('Z_Flete', 'Z_Litros', 'Z_CxL'):
                    if _zcol in df_m_view.columns:
                        styled_m = styled_m.map(colorear_z_celda, subset=[_zcol])
                st.dataframe(styled_m, use_container_width=True, height=380)

                xlsx_m = styled_to_xlsx_bytes(
                    styled_m,
                    sheet_name='Anomalías MEDIO',
                    column_formats={
                        'Flete': '"$"#,##0.00',
                        'Litros Fact': '#,##0',
                        'Costo por Litro': '"$"#,##0.00',
                        'Z_Flete': '+0.00;-0.00;0.00',
                        'Z_Litros': '+0.00;-0.00;0.00',
                        'Z_CxL': '+0.00;-0.00;0.00',
                    }
                )
                st.download_button(
                    "📥 Descargar registros MEDIO (Excel con formato)",
                    xlsx_m, "anomalias_medio.xlsx", XLSX_MIME,
                    key="dl_kpi_medio"
                )
            else:
                st.success("✅ No hay anomalías de severidad MEDIO con los filtros actuales.")

    # ╔══════════════════════════════════════════════════════════╗
# ║  PESTAÑA 2 — FACTURAS SIN MATCH EN SAP                   ║
# ╚══════════════════════════════════════════════════════════╝
with tab2:
    if not (df_log_raw is not None and df_sap is not None):
        st.info("📤 Sube el archivo Excel desde la barra lateral para activar esta pestaña.")

if df_log_raw is not None and df_sap is not None:
    with tab2:
        st.markdown("**Conciliación entre facturas de transporte y SAP** — detecta errores de captura, "
                    "facturas faltantes y discrepancias de monto.")

        # Alerta explicativa
        st.markdown("""
        <div style="
            background: linear-gradient(135deg, #DBEAFE 0%, #BFDBFE 100%);
            border-left: 6px solid #1E3A8A;
            padding: 16px 20px;
            border-radius: 12px;
            margin: 16px 0;
        ">
            <div style="font-size: 16px; font-weight: 600; color: #1E3A8A; margin-bottom: 8px;">
                ℹ️ Cómo se hace la conciliación
            </div>
            <div style="color: #1E3A8A; line-height: 1.6;">
                Para cada factura de transporte registrada en <em>Logística Nac</em>, el sistema busca su 
                <strong>match exacto en SAP proveedores</strong>. Si no la encuentra, intenta resolver 
                mediante similitud textual (errores de captura) y por monto. Cada registro recibe una 
                categoría que indica la naturaleza de la discrepancia y una explicación humana de la causa.
                <br><br>
                Las filas con <strong>proveedor "CLIENTE"</strong> y las que tienen la columna 
                <em>Factura transporte</em> vacía se omiten automáticamente.
            </div>
        </div>
        """, unsafe_allow_html=True)

        # Sidebar específico de tab2
        st.sidebar.markdown("---")
        st.sidebar.markdown("### 🔎 Pestaña: Facturas vs SAP")
        tolerancia = st.sidebar.number_input(
            "Tolerancia de monto ($)", min_value=0.0, max_value=100.0, value=1.0, step=0.5,
            help="Diferencia máxima en pesos para considerar que un monto coincide"
        )

        # Ejecutar auditoría
        with st.spinner("🔄 Cruzando facturas con SAP..."):
            df_audit = auditar_facturas(df_log_raw, df_sap, tolerancia_monto=tolerancia)

        if len(df_audit) == 0:
            st.warning("No se encontraron registros válidos para auditar.")
        else:
            # KPIs de la pestaña 2
            n_total = len(df_audit)
            n_ok = (df_audit['Categoría'] == 'COINCIDE_EXACTO').sum()
            n_provision = (df_audit['Categoría'] == 'PROVISION').sum()
            n_problemas = n_total - n_ok - n_provision
            n_typos = df_audit['Categoría'].isin(
                ['TYPO_CON_MONTO_COINCIDENTE', 'UUID_TYPO_CON_MONTO_COINCIDENTE']
            ).sum()

            # ─── KPIs CLICKABLES (v16: mutuamente exclusivos) ──
            if 'kpi2_active' not in st.session_state:
                st.session_state.kpi2_active = None

            def _toggle_kpi2(name):
                st.session_state.kpi2_active = (
                    None if st.session_state.kpi2_active == name else name
                )

            kpi_b1, kpi_b2, kpi_b3, kpi_b4 = st.columns(4)
            with kpi_b1:
                _arr = "▲" if st.session_state.kpi2_active == 'auditadas' else "▼"
                if st.button(
                    f"📊 Facturas auditadas\n\n{n_total:,}  {_arr}",
                    key="kpi2_btn_auditadas",
                    use_container_width=True,
                    help="Click para ver el resumen completo de todas las facturas"
                ):
                    _toggle_kpi2('auditadas')
            with kpi_b2:
                _arr = "▲" if st.session_state.kpi2_active == 'match' else "▼"
                if st.button(
                    f"✅ Match exacto\n\n{n_ok:,}  {_arr}",
                    key="kpi2_btn_match",
                    use_container_width=True,
                    help="Click para ver solo las facturas con coincidencia exacta en SAP"
                ):
                    _toggle_kpi2('match')
            with kpi_b3:
                _arr = "▲" if st.session_state.kpi2_active == 'discrepancias' else "▼"
                if st.button(
                    f"⚠️ Con discrepancias\n\n{n_problemas:,}  {_arr}",
                    key="kpi2_btn_discrepancias",
                    use_container_width=True,
                    help="Click para ver solo las facturas con discrepancias"
                ):
                    _toggle_kpi2('discrepancias')
            with kpi_b4:
                _arr = "▲" if st.session_state.kpi2_active == 'typos' else "▼"
                if st.button(
                    f"🟢 Errores de captura confirmados\n\n{n_typos:,}  {_arr}",
                    key="kpi2_btn_typos",
                    use_container_width=True,
                    help="Click para ver errores de captura con monto coincidente"
                ):
                    _toggle_kpi2('typos')

            # ─── CSS dinámico: desvanecer botones inactivos ──
            _active2 = st.session_state.kpi2_active
            if _active2 is not None:
                _all2 = ('auditadas', 'match', 'discrepancias', 'typos')
                _inactive2 = [k for k in _all2 if k != _active2]
                _fade_css2 = ""
                for _k in _inactive2:
                    _fade_css2 += (
                        f'[class*="st-key-kpi2_btn_{_k}"] button,'
                        f'body [class*="st-key-kpi2_btn_{_k}"] button {{'
                        f'  opacity: 0.40 !important;'
                        f'  background: linear-gradient(135deg, #94A3B8 0%, #64748B 100%) !important;'
                        f'  box-shadow: none !important;'
                        f'  transform: none !important;'
                        f'}}'
                        f'[class*="st-key-kpi2_btn_{_k}"] button:hover,'
                        f'body [class*="st-key-kpi2_btn_{_k}"] button:hover {{'
                        f'  opacity: 0.70 !important;'
                        f'  background: linear-gradient(135deg, #64748B 0%, #475569 100%) !important;'
                        f'  transform: none !important;'
                        f'  box-shadow: 0 2px 6px rgba(100, 116, 139, 0.25) !important;'
                        f'}}'
                        f'[class*="st-key-kpi2_btn_{_k}"] button p,'
                        f'body [class*="st-key-kpi2_btn_{_k}"] button p {{'
                        f'  color: #E2E8F0 !important;'
                        f'}}'
                    )
                st.markdown(f"<style>{_fade_css2}</style>", unsafe_allow_html=True)

            # ─── PANEL: contenido condicional según botón activo ─
            if _active2 is None:
                st.info("👆 Selecciona uno de los botones de arriba para ver el detalle. "
                        "**Facturas auditadas** muestra todo; los otros filtran a un subconjunto.")
            else:
                # Pre-filtrar df_audit según el botón activo
                if _active2 == 'auditadas':
                    df_audit_view = df_audit.copy()
                    panel_titulo = "📊 Vista completa: todas las facturas auditadas"
                    panel_clase = ""
                elif _active2 == 'match':
                    df_audit_view = df_audit[df_audit['Categoría'] == 'COINCIDE_EXACTO'].copy()
                    panel_titulo = "✅ Vista filtrada: solo facturas con match exacto en SAP"
                    panel_clase = ""
                elif _active2 == 'discrepancias':
                    df_audit_view = df_audit[
                        ~df_audit['Categoría'].isin(['COINCIDE_EXACTO', 'PROVISION'])
                    ].copy()
                    panel_titulo = "⚠️ Vista filtrada: solo facturas con discrepancias"
                    panel_clase = "alto"
                else:  # 'typos'
                    df_audit_view = df_audit[df_audit['Categoría'].isin(
                        ['TYPO_CON_MONTO_COINCIDENTE', 'UUID_TYPO_CON_MONTO_COINCIDENTE']
                    )].copy()
                    panel_titulo = "🟢 Vista filtrada: errores de captura con alta confianza"
                    panel_clase = "clientes"

                st.markdown(
                    f'<div class="kpi-detail-panel {panel_clase}"><h4>{panel_titulo}</h4></div>',
                    unsafe_allow_html=True
                )

                if len(df_audit_view) == 0:
                    st.info("Sin registros en este subconjunto.")

            st.markdown("---")

            if _active2 is not None:
                # Resumen por categoría
                st.subheader("📋 Resumen por categoría")
                resumen = df_audit_view['Categoría'].value_counts().reset_index()
                resumen.columns = ['Categoría', 'Cantidad']
                resumen['Etiqueta'] = resumen['Categoría'].map(
                    lambda c: CATEGORIAS_INFO.get(c, {}).get('label', c)
                )
                resumen['Descripción'] = resumen['Categoría'].map(
                    lambda c: CATEGORIAS_INFO.get(c, {}).get('desc', '')
                )

                # Mostrar como cards
                for _, row in resumen.iterrows():
                    info = CATEGORIAS_INFO.get(row['Categoría'], {})
                    color = info.get('color', '#475569')
                    bg = info.get('bg', '#F1F5F9')
                    label = info.get('label', row['Categoría'])
                    desc = info.get('desc', '')
                    st.markdown(f"""
                    <div style="
                        background: {bg};
                        border-left: 5px solid {color};
                        padding: 12px 16px;
                        border-radius: 8px;
                        margin: 6px 0;
                        display: flex;
                        justify-content: space-between;
                        align-items: center;
                    ">
                        <div style="flex: 1;">
                            <div style="font-weight: 600; color: {color}; font-size: 15px;">
                                {label}
                            </div>
                            <div style="color: #475569; font-size: 12px; margin-top: 4px;">
                                {desc}
                            </div>
                        </div>
                        <div style="
                            background: {color}; color: white; 
                            padding: 8px 16px; border-radius: 8px;
                            font-weight: 700; font-size: 18px; min-width: 60px; text-align: center;
                        ">
                            {row['Cantidad']}
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                st.markdown("---")

                # ─── Filtros de tabla — título y filtros condicionales según panel ─
                if _active2 == 'match':
                    st.subheader("🔍 Detalle de registros con match exacto")
                elif _active2 == 'typos':
                    st.subheader("🔍 Detalle de registros con errores de captura")
                else:  # 'auditadas' o 'discrepancias'
                    st.subheader("🔍 Detalle de registros con discrepancias")

                if _active2 == 'auditadas':
                    # Vista completa: las 3 columnas con cat_sel, prov_sel y ocultar_ok
                    col_f1, col_f2, col_f3 = st.columns([2, 2, 1])
                    with col_f1:
                        cat_disp = sorted([c for c in df_audit_view['Categoría'].unique()
                                           if c not in ['COINCIDE_EXACTO', 'PROVISION']])
                        cat_sel = st.multiselect(
                            "Filtrar por categoría",
                            options=cat_disp,
                            default=cat_disp,
                            format_func=lambda c: CATEGORIAS_INFO.get(c, {}).get('label', c),
                            key="cat_sel_t2_auditadas"
                        )
                    with col_f2:
                        prov_uniq = sorted(df_audit_view['Proveedor transporte'].dropna().unique())
                        prov_sel = st.multiselect("Filtrar por proveedor",
                                                   options=prov_uniq, default=[],
                                                   key="prov_sel_t2_auditadas")
                    with col_f3:
                        ocultar_ok = st.checkbox("Ocultar matches OK", value=True,
                                                  key="ocultar_ok_t2")
                elif _active2 == 'match':
                    # Match exacto: sin checkbox de ocultar; solo proveedor (categoría única)
                    col_f1, col_f2 = st.columns([1, 1])
                    with col_f1:
                        prov_uniq = sorted(df_audit_view['Proveedor transporte'].dropna().unique())
                        prov_sel = st.multiselect("Filtrar por proveedor",
                                                   options=prov_uniq, default=[],
                                                   key="prov_sel_t2_match")
                    with col_f2:
                        st.empty()  # placeholder
                    cat_sel = []
                    ocultar_ok = False
                else:
                    # Discrepancias o typos: sin checkbox; con cat_sel + prov_sel
                    col_f1, col_f2 = st.columns([1, 1])
                    with col_f1:
                        cat_disp = sorted(df_audit_view['Categoría'].unique().tolist())
                        cat_sel = st.multiselect(
                            "Filtrar por categoría",
                            options=cat_disp,
                            default=cat_disp,
                            format_func=lambda c: CATEGORIAS_INFO.get(c, {}).get('label', c),
                            key=f"cat_sel_t2_{_active2}"
                        )
                    with col_f2:
                        prov_uniq = sorted(df_audit_view['Proveedor transporte'].dropna().unique())
                        prov_sel = st.multiselect("Filtrar por proveedor",
                                                   options=prov_uniq, default=[],
                                                   key=f"prov_sel_t2_{_active2}")
                    ocultar_ok = False

                df_show2 = df_audit_view.copy()
                if ocultar_ok:
                    df_show2 = df_show2[~df_show2['Categoría'].isin(['COINCIDE_EXACTO', 'PROVISION'])]
                if cat_sel:
                    df_show2 = df_show2[df_show2['Categoría'].isin(cat_sel)]
                if prov_sel:
                    df_show2 = df_show2[df_show2['Proveedor transporte'].isin(prov_sel)]

                st.info(f"📌 Mostrando **{len(df_show2)}** registros")

                # Renombrar categoría para usuario
                df_show2 = df_show2.copy()
                df_show2['Categoría (humana)'] = df_show2['Categoría'].map(
                    lambda c: CATEGORIAS_INFO.get(c, {}).get('label', c)
                )

                # Columnas a mostrar
                cols_audit = [
                    'Fecha Factura', 'Nombre de Cliente', 'Remisión',
                    'factura_evaluada', 'Proveedor transporte',
                    'Total Flete ', 'Total Flete',
                    'Categoría (humana)', 'Razón de No Coincidencia'
                ]
                cols_audit = [c for c in cols_audit if c in df_show2.columns]
                df_view = df_show2[cols_audit].copy()
                if 'Fecha Factura' in df_view.columns:
                    df_view['Fecha Factura'] = pd.to_datetime(df_view['Fecha Factura'], errors='coerce').dt.strftime('%d/%m/%Y')


                def colorear_cat(val):
                    for k, info in CATEGORIAS_INFO.items():
                        if info['label'] == val:
                            return f'background-color: {info["bg"]}; color: {info["color"]}; font-weight: 600;'
                    return ''


                fmt2 = {}
                if 'Total Flete' in df_view.columns: fmt2['Total Flete'] = '${:,.2f}'
                if 'Total Flete ' in df_view.columns: fmt2['Total Flete '] = '${:,.2f}'

                styled2 = df_view.style.format(fmt2)
                if 'Categoría (humana)' in df_view.columns:
                    styled2 = styled2.map(colorear_cat, subset=['Categoría (humana)'])

                st.dataframe(styled2, use_container_width=True, height=500)

                # Descarga Excel con formato preservado (colores por categoría)
                xlsx2 = styled_to_xlsx_bytes(
                    styled2,
                    sheet_name='Auditoría SAP',
                    column_formats={
                        'Total Flete': '"$"#,##0.00',
                        'Total Flete ': '"$"#,##0.00',
                    }
                )
                st.download_button(
                    "📥 Descargar resultados (Excel con formato)",
                    xlsx2, "auditoria_facturas_sap.xlsx", XLSX_MIME,
                    key="dl_t2"
                )

                # ─── Diagnósticos detallados — solo si NO es 'match' ─
                if _active2 != 'match':
                    st.markdown("---")
                    st.subheader("🔍 Casos a revisar prioritariamente")

                    # Categorías mostradas según el panel activo
                    if _active2 == 'auditadas':
                        # Todo lo que no es match/provisión/gris: amarillas + naranjas + rojas
                        cats_criticas = [
                            'MONTO_COINCIDE_REF_DISTINTA', 'UUID_MONTO_COINCIDE_DISTINTO',  # amarillas
                            'POSIBLE_TYPO_SIN_MONTO', 'UUID_POSIBLE_TYPO_SIN_MONTO',         # naranjas
                            'NO_ENCONTRADO', 'UUID_NO_ENCONTRADO',                            # rojas
                        ]
                        sub_text = "**Registros que requieren revisión** (categorías 🟡 amarillas, 🟠 naranjas y 🔴 rojas):"
                    elif _active2 == 'discrepancias':
                        # Solo rojas + naranjas
                        cats_criticas = [
                            'POSIBLE_TYPO_SIN_MONTO', 'UUID_POSIBLE_TYPO_SIN_MONTO',  # naranjas
                            'NO_ENCONTRADO', 'UUID_NO_ENCONTRADO',                     # rojas
                        ]
                        sub_text = "**Casos prioritarios** (🔴 rojos y 🟠 naranjas — necesitan acción):"
                    else:  # 'typos'
                        cats_criticas = ['TYPO_CON_MONTO_COINCIDENTE', 'UUID_TYPO_CON_MONTO_COINCIDENTE']
                        sub_text = "**Errores de captura con alta confianza** (monto coincide exactamente):"

                    criticos = df_audit_view[df_audit_view['Categoría'].isin(cats_criticas)].head(20)

                    if len(criticos) > 0:
                        st.markdown(sub_text)
                        for _, row in criticos.iterrows():
                            info = CATEGORIAS_INFO.get(row['Categoría'], {})
                            fecha_str = pd.to_datetime(row['Fecha Factura'], errors='coerce')
                            fecha_str = fecha_str.strftime("%d/%m/%Y") if pd.notna(fecha_str) else "N/A"
                            st.markdown(f"""
                            <div style="
                                background: {info.get('bg', '#F0FDF4')};
                                border-left: 4px solid {info.get('color', '#16A34A')};
                                padding: 12px 16px;
                                border-radius: 8px;
                                margin: 6px 0;
                            ">
                                <div style="font-weight: 600; color: {info.get('color', '#16A34A')};">
                                    {row.get('Nombre de Cliente', 'N/A')} · {fecha_str} ·
                                    Remisión {row.get('Remisión', 'N/A')}
                                </div>
                                <div style="color: #1E293B; font-size: 13px; margin-top: 4px; line-height: 1.5;">
                                    {row.get('Razón de No Coincidencia', '')}
                                </div>
                            </div>
                            """, unsafe_allow_html=True)
                    else:
                        st.success("✅ No se detectaron casos prioritarios en este subconjunto.")

    # ╔══════════════════════════════════════════════════════════╗
# ║  PESTAÑA 3 — DIAGNÓSTICO DE VENTAS (Pipeline LLMX)        ║
# ╚══════════════════════════════════════════════════════════╝
with tab3:
    # ─── Estética dark/neón replicando la app original ──────────
    # Paleta: BG_DARK #0F1117, BG_CARD #1A1D27, BG_INPUT #12141C,
    # ACCENT #00E5A0 (verde menta), ACCENT2 #0099FF (azul eléctrico),
    # SUCCESS #00C97A, ERROR #FF4757, WARN #FF6B35
    st.markdown("""
    <style>
        .pipe-wrap {
            background: #0F1117;
            border-radius: 16px;
            padding: 24px 28px;
            margin-top: 8px;
            margin-bottom: 8px;
            font-family: 'Courier New', Courier, monospace;
            color: #E8EAF0;
        }
        .pipe-wrap h2 {
            color: #00E5A0 !important;
            border-bottom: 1px solid #00E5A0;
            padding-bottom: 6px;
            font-family: 'Courier New', Courier, monospace;
            letter-spacing: 1px;
        }
        .pipe-wrap .subtitle {
            color: #5A5F72;
            font-size: 12px;
            letter-spacing: 1px;
            margin-bottom: 20px;
            text-transform: uppercase;
        }
        .pipe-wrap .section-label {
            color: #5A5F72;
            font-size: 10px;
            letter-spacing: 2px;
            font-weight: 700;
            text-transform: uppercase;
            margin: 18px 0 8px 0;
        }
        .pipe-card {
            background: #1A1D27;
            border: 1px solid #2A2D3A;
            border-radius: 10px;
            padding: 16px;
            text-align: center;
        }
        .pipe-card .icon-sales { color: #00E5A0; font-size: 28px; }
        .pipe-card .icon-llmx  { color: #0099FF; font-size: 28px; }
        .pipe-card .title  {
            color: #E8EAF0; font-weight: 700; margin: 6px 0 2px 0;
            font-family: 'Courier New', Courier, monospace;
        }
        .pipe-card .subtitle-card {
            color: #5A5F72; font-size: 11px;
            font-family: 'Courier New', Courier, monospace;
        }
        .pipe-card .filename-ok    { color: #00E5A0; font-size: 11px; margin-top: 8px; }
        .pipe-card .filename-ok2   { color: #0099FF; font-size: 11px; margin-top: 8px; }
        .pipe-card .filename-warn  { color: #FF6B35; font-size: 11px; margin-top: 8px; }
        .pipe-card .filename-empty { color: #5A5F72; font-size: 11px; margin-top: 8px; font-style: italic; }
        .pipe-console {
            background: #12141C;
            border: 1px solid #2A2D3A;
            border-radius: 10px;
            padding: 14px 18px;
            font-family: 'Courier New', Courier, monospace;
            font-size: 12.5px;
            line-height: 1.55;
            color: #E8EAF0;
            white-space: pre-wrap;
            max-height: 540px;
            overflow-y: auto;
        }
        .pipe-console .t-sep    { color: #5A5F72; }
        .pipe-console .t-titulo { color: #00E5A0; font-weight: 700; }
        .pipe-console .t-ok     { color: #00C97A; }
        .pipe-console .t-error  { color: #FF4757; }
        .pipe-console .t-warn   { color: #FF6B35; }
        .pipe-console .t-alerta { color: #FF6B35; font-weight: 700; }
        .pipe-console .t-info   { color: #E8EAF0; }
        .pipe-console .t-header { color: #0099FF; font-weight: 700; }
        .pipe-console .t-fila   { color: #E8EAF0; }
        .pipe-console .t-total  { color: #00E5A0; font-weight: 700; }
    </style>
    """, unsafe_allow_html=True)

    # Título estilo app original
    st.markdown(
        '<div class="pipe-wrap" style="padding-bottom: 4px;">'
        '<h2>DIAGNÓSTICO DE VENTAS</h2>'
        '<div class="subtitle">Pipeline automatizado · Limpieza · Análisis · Investigación</div>'
        '</div>',
        unsafe_allow_html=True
    )

    # Estado persistente
    if 'pipe_sale_bytes' not in st.session_state:
        st.session_state.pipe_sale_bytes = None
        st.session_state.pipe_sale_name = None
        st.session_state.pipe_sale_tipo = None
    if 'pipe_pl_bytes' not in st.session_state:
        st.session_state.pipe_pl_bytes = None
        st.session_state.pipe_pl_name = None
        st.session_state.pipe_pl_tipo = None
    if 'pipe_log' not in st.session_state:
        st.session_state.pipe_log = []
    if 'pipe_resultado' not in st.session_state:
        st.session_state.pipe_resultado = None

    # ─── ARCHIVOS DE ENTRADA (dos drop zones) ────────────────────
    st.markdown(
        '<div class="pipe-wrap"><div class="section-label">▼ Archivos de entrada</div>',
        unsafe_allow_html=True
    )

    col_sales, col_llmx = st.columns(2)

    with col_sales:
        archivo_sales = st.file_uploader(
            "Sales Data Base (Monthly) — Sale_Database · Nacionales · Extranjeros",
            type=['xlsx', 'xlsm'],
            key='upload_sales_pipe',
            help="Excel con hojas Sale_Database, Nacionales y Extranjeros"
        )
        if archivo_sales is not None:
            data = archivo_sales.getvalue()
            tipo = detectar_tipo_archivo(data)
            st.session_state.pipe_sale_bytes = data
            st.session_state.pipe_sale_name = archivo_sales.name
            st.session_state.pipe_sale_tipo = tipo
            if tipo == 'sales':
                clase = 'filename-ok'
                etiqueta = f"✔  {archivo_sales.name}"
            elif tipo == 'llmx':
                clase = 'filename-warn'
                etiqueta = f"⚠  {archivo_sales.name} parece ser un archivo LLMX"
            else:
                clase = 'filename-warn'
                etiqueta = f"⚠  {archivo_sales.name} no contiene las hojas esperadas"
        else:
            clase = 'filename-empty'
            etiqueta = "Sin archivo cargado — sube el Sales Data Base"
        st.markdown(
            f'<div class="pipe-card">'
            f'<div class="icon-sales">⬇</div>'
            f'<div class="title">SALES DATA BASE</div>'
            f'<div class="subtitle-card">Sale_Database · Nacionales · Extranjeros</div>'
            f'<div class="{clase}">{etiqueta}</div>'
            f'</div>',
            unsafe_allow_html=True
        )

    with col_llmx:
        archivo_llmx = st.file_uploader(
            "LLMX_YYYY — P&L · Tabs mensuales N_YY",
            type=['xlsx', 'xlsm'],
            key='upload_llmx_pipe',
            help="Excel LLMX con hoja P&L YY y tabs mensuales N_YY"
        )
        if archivo_llmx is not None:
            data = archivo_llmx.getvalue()
            tipo = detectar_tipo_archivo(data)
            st.session_state.pipe_pl_bytes = data
            st.session_state.pipe_pl_name = archivo_llmx.name
            st.session_state.pipe_pl_tipo = tipo
            if tipo == 'llmx':
                clase = 'filename-ok2'
                etiqueta = f"✔  {archivo_llmx.name}"
            elif tipo == 'sales':
                clase = 'filename-warn'
                etiqueta = f"⚠  {archivo_llmx.name} parece ser un Sales Data Base"
            else:
                clase = 'filename-warn'
                etiqueta = f"⚠  {archivo_llmx.name} no contiene las hojas esperadas"
        else:
            clase = 'filename-empty'
            etiqueta = "Sin archivo cargado — sube el LLMX"
        st.markdown(
            f'<div class="pipe-card">'
            f'<div class="icon-llmx">⬇</div>'
            f'<div class="title">LLMX_YYYY</div>'
            f'<div class="subtitle-card">P&L · Tabs mensuales N_YY</div>'
            f'<div class="{clase}">{etiqueta}</div>'
            f'</div>',
            unsafe_allow_html=True
        )

    st.markdown('</div>', unsafe_allow_html=True)

    # ─── BOTÓN EJECUTAR ─────────────────────────────────────────
    sale_ok = (st.session_state.pipe_sale_bytes is not None
               and st.session_state.pipe_sale_tipo == 'sales')
    pl_ok = (st.session_state.pipe_pl_bytes is not None
             and st.session_state.pipe_pl_tipo == 'llmx')

    # Botón estilo neón verde — Streamlit no permite estilizar botones por id,
    # pero podemos estilar todos los buttons del documento. En su lugar,
    # usamos un botón estándar con el estilo Primary y un caption explicativo.
    st.markdown("""
    <style>
        div.stButton > button[kind="primary"] {
            background: #00E5A0 !important;
            color: #0F1117 !important;
            border: none !important;
            font-family: 'Courier New', Courier, monospace !important;
            font-weight: 700 !important;
            letter-spacing: 1px !important;
            padding: 10px 28px !important;
            border-radius: 8px !important;
        }
        div.stButton > button[kind="primary"]:hover {
            background: #00FFB2 !important;
            color: #0F1117 !important;
        }
    </style>
    """, unsafe_allow_html=True)

    btn_col1, btn_col2, btn_col3 = st.columns([1, 1, 1])
    with btn_col2:
        ejecutar = st.button(
            "▶   EJECUTAR PIPELINE",
            type="primary",
            disabled=not (sale_ok and pl_ok),
            use_container_width=True,
            key="btn_run_pipe"
        )

    if not (sale_ok and pl_ok):
        faltantes = []
        if not sale_ok:
            faltantes.append("Sales Data Base")
        if not pl_ok:
            faltantes.append("LLMX")
        st.markdown(
            f'<div style="text-align:center; color:#FF6B35; '
            f'font-family:Courier New,monospace; font-size:11px; margin-top:6px;">'
            f'⚠  Carga ambos archivos correctamente antes de ejecutar '
            f'(falta: {", ".join(faltantes)}).</div>',
            unsafe_allow_html=True
        )

    # ─── EJECUCIÓN ──────────────────────────────────────────────
    if ejecutar:
        st.session_state.pipe_log = []

        def _log(msg, tag="info"):
            st.session_state.pipe_log.append((msg, tag))

        try:
            with st.spinner("Ejecutando pipeline..."):
                resultado = run_pipeline_streamlit(
                    st.session_state.pipe_sale_bytes,
                    st.session_state.pipe_sale_name,
                    st.session_state.pipe_pl_bytes,
                    st.session_state.pipe_pl_name,
                    _log,
                )
            _log(f"\n{SEP_PIPE}", "sep")
            _log("  PIPELINE FINALIZADO CORRECTAMENTE", "ok")
            _log(SEP_PIPE, "sep")
            st.session_state.pipe_resultado = resultado
        except FormatError as e:
            _log(f"\n{'!' * 60}", "error")
            _log(f"  ⚠  ALERTA DE FORMATO", "error")
            _log(f"{'!' * 60}", "error")
            _log(f"  {str(e)}", "warn")
            for d in e.detalles:
                _log(f"    → {d}", "warn")
            _log(f"{'!' * 60}", "error")
            st.session_state.pipe_resultado = None
        except Exception as e:
            _log(f"\n  ERROR INESPERADO: {str(e)}", "error")
            _log(traceback.format_exc(), "error")
            st.session_state.pipe_resultado = None

    # ─── CONSOLA DE SALIDA ──────────────────────────────────────
    st.markdown(
        '<div class="pipe-wrap"><div class="section-label">▼ Consola de salida</div>',
        unsafe_allow_html=True
    )

    if not st.session_state.pipe_log:
        contenido = (
            '<span class="t-sep">' + SEP_PIPE + '</span>\n'
            '<span class="t-titulo">  PIPELINE COMPLETO — DIAGNÓSTICO DE VENTAS</span>\n'
            '<span class="t-sep">' + SEP_PIPE + '</span>\n'
            '<span class="t-info">  Carga los dos archivos Excel y presiona Ejecutar.</span>\n'
        )
    else:
        partes = []
        for msg, tag in st.session_state.pipe_log:
            # Escapar HTML básico
            msg_safe = (str(msg)
                        .replace('&', '&amp;')
                        .replace('<', '&lt;')
                        .replace('>', '&gt;'))
            partes.append(f'<span class="t-{tag}">{msg_safe}</span>')
        contenido = '\n'.join(partes)

    st.markdown(f'<div class="pipe-console">{contenido}</div>',
                unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)

    # ─── DESCARGA DE ARCHIVO LIMPIO ─────────────────────────────
    if (st.session_state.pipe_resultado
            and st.session_state.pipe_resultado.get('cleaned_sale_bytes')):
        nombre_original = st.session_state.pipe_sale_name or 'Sales_Data_Base.xlsx'
        base, ext = os.path.splitext(nombre_original)
        ext = ext or '.xlsx'
        nombre_limpio = f"{base}_LIMPIO{ext}"
        st.markdown(
            '<div class="pipe-wrap"><div class="section-label">▼ Resultado</div>',
            unsafe_allow_html=True
        )
        st.download_button(
            label=f"⬇  Descargar Sales Data Base limpio  ({nombre_limpio})",
            data=st.session_state.pipe_resultado['cleaned_sale_bytes'],
            file_name=nombre_limpio,
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            key='dl_sale_clean',
            use_container_width=True,
        )
        st.markdown('</div>', unsafe_allow_html=True)


# ============================================================
# FOOTER
# ============================================================
st.markdown("---")
st.caption(
    "💡 **Dashboard de Auditoría Logística** · "
    "Análisis estadístico de fletes (Z-scores por cliente) + "
    "Conciliación SAP con detección de errores de captura + "
    "Diagnóstico de Ventas (Pipeline LLMX vs Sale_Database)"
)

# [v9-restructured]

# [v12-stripped-legacy]

# [v16-tab2-wrapped]
